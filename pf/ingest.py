from __future__ import annotations

from dataclasses import dataclass
from datetime import date as today_date
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from pf.config import CardConfig
from pf.db import (
    SyncResult,
    add_pending_review,
    count_pending_reviews,
    bulk_update_categories_by_row_hash,
    is_imported,
    insert_transactions,
    now_iso,
    register_import,
    sync_transactions_by_row_hash,
    upsert_credit_card_transactions,
)
from pf.excel_unified import read_contas_casa_sheet, read_debitos_sheet, read_receitas_sheet
from pf.importers import import_credit_card_csv, import_debit_xlsx, import_income_xlsx
from pf.importers.credit_card_csv import extract_statement_due_date_from_path
from pf.importers.credit_card_categories_xlsx import read_credit_card_categories_xlsx
from pf.importers.credit_card_master_xlsx import read_credit_card_master_xlsx
from pf.rules_engine import apply_rules_to_rows
from pf.utils import month_add, normalize_str, sha256_file, sha256_text


ImporterKind = Literal["credit_card_csv", "debit_xlsx", "income_xlsx"]


@dataclass(frozen=True)
class IngestResult:
    imported: bool
    rows_read: int
    rows_inserted: int
    file_hash: str


@dataclass(frozen=True)
class UnifiedSyncResult:
    credit_card: SyncResult
    debit: SyncResult
    income: SyncResult
    household: SyncResult


def _workbook_has_sheet(path: Path, sheet_name: str) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return False
    return sheet_name in wb.sheetnames


def ingest_credit_card_csv(
    conn,
    *,
    path: Path,
    card: CardConfig,
    person: str | None = None,
    rules: list[dict] | None = None,
    force: bool = False,
) -> IngestResult:
    file_hash = sha256_file(path)
    # For open statements, always re-process the CSV until the due date has passed.
    # This allows users to re-download/overwrite the same CSV while the statement is still changing.
    due_dt = extract_statement_due_date_from_path(path)
    is_open_statement = due_dt is not None and today_date.today() <= due_dt
    if not force and not is_open_statement and is_imported(conn, file_hash):
        return IngestResult(imported=False, rows_read=0, rows_inserted=0, file_hash=file_hash)
    rows = import_credit_card_csv(
        path,
        card=card,
        source_hash=file_hash,
        source_file=str(path),
        person=person,
    )
    if rules:
        apply_rules_to_rows(rows, rules)
    inserted = upsert_credit_card_transactions(conn, rows)

    # Statement CSVs whose filename carries a due date are treated as the
    # source-of-truth snapshot for that (card, due_date) statement.
    # Any currently visible row from the same statement that is not present
    # in this import is soft-hidden to avoid stale + new rows being summed together.
    due_dt = extract_statement_due_date_from_path(path)
    if due_dt is not None:
        imported_hashes = sorted(
            {
                str(r.get("row_hash") or "").strip()
                for r in rows
                if str(r.get("row_hash") or "").strip()
            }
        )
        if imported_hashes:
            placeholders = ", ".join(["?"] * len(imported_hashes))
            params = [now_iso(), card.name, due_dt.isoformat(), *imported_hashes]
            conn.execute(
                f"""
                UPDATE transactions
                SET hidden_in_excel = 1, updated_at = ?
                WHERE payment_method = 'credit_card'
                  AND account = ?
                  AND statement_due_date = ?
                  AND COALESCE(hidden_in_excel, 0) = 0
                  AND row_hash NOT IN ({placeholders})
                """,
                params,
            )
            conn.commit()

    register_import(conn, file_hash=file_hash, file_path=str(path), importer=f"credit_card_csv:{card.id}", rows=len(rows))
    return IngestResult(imported=True, rows_read=len(rows), rows_inserted=inserted, file_hash=file_hash)


def ingest_debit_xlsx(conn, *, path: Path, rules: list[dict] | None = None, force: bool = False) -> IngestResult:
    file_hash = sha256_file(path)
    if not force and is_imported(conn, file_hash):
        return IngestResult(imported=False, rows_read=0, rows_inserted=0, file_hash=file_hash)
    rows = import_debit_xlsx(path, source_hash=file_hash, source_file=str(path))
    if rules:
        apply_rules_to_rows(rows, rules)
    inserted = insert_transactions(conn, rows)
    register_import(conn, file_hash=file_hash, file_path=str(path), importer="debit_xlsx", rows=len(rows))
    return IngestResult(imported=True, rows_read=len(rows), rows_inserted=inserted, file_hash=file_hash)


def ingest_income_xlsx(conn, *, path: Path, rules: list[dict] | None = None, force: bool = False) -> IngestResult:
    file_hash = sha256_file(path)
    if not force and is_imported(conn, file_hash):
        return IngestResult(imported=False, rows_read=0, rows_inserted=0, file_hash=file_hash)
    rows = import_income_xlsx(path, source_hash=file_hash, source_file=str(path))
    if rules:
        apply_rules_to_rows(rows, rules)
    inserted = insert_transactions(conn, rows)
    register_import(conn, file_hash=file_hash, file_path=str(path), importer="income_xlsx", rows=len(rows))
    return IngestResult(imported=True, rows_read=len(rows), rows_inserted=inserted, file_hash=file_hash)


def scan_raw_data(base_dir: Path) -> list[Path]:
    raw = base_dir / "raw_data"
    if not raw.exists():
        return []
    files: list[Path] = []
    for ext in ("*.csv", "*.CSV", "*.xlsx", "*.XLSX"):
        files.extend(raw.rglob(ext))

    def _is_ignored(p: Path) -> bool:
        # Convention: `raw_data/old/` is an archive folder and should never be scanned.
        # (It exists so users can keep closed/imported statements without reprocessing.)
        parts = [part.lower() for part in p.parts]
        try:
            raw_idx = parts.index("raw_data")
        except ValueError:
            return False
        return (raw_idx + 1) < len(parts) and parts[raw_idx + 1] == "old"

    return sorted([p for p in files if p.is_file() and not _is_ignored(p)])


def apply_credit_card_categories_from_xlsx(
    conn,
    *,
    path: Path,
    allow_clear: bool = False,
) -> tuple[int, int]:
    """
    Updates the DB based on user edits in `templates/cartao_credito.xlsx`.

    Returns `(updated_count, missing_count)`, where `missing_count` are rows
    whose `row_hash` wasn't found in the DB.
    """
    updates = read_credit_card_categories_xlsx(path)
    return bulk_update_categories_by_row_hash(conn, updates, allow_clear=allow_clear)


def apply_debit_categories_from_xlsx(
    conn,
    *,
    path: Path,
    allow_clear: bool = False,
) -> tuple[int, int]:
    """
    Updates the DB based on user edits in `raw_data/debitos.xlsx`.

    This allows treating the debit input workbook as the source of truth for
    categorization (without re-inserting duplicated rows).
    """
    rows = import_debit_xlsx(path, source_hash="apply_debit_xlsx", source_file=str(path))
    updates = [
        {
            "row_hash": r.get("row_hash"),
            "category": r.get("category"),
            "subcategory": r.get("subcategory"),
            "reimbursable": r.get("reimbursable"),
        }
        for r in rows
        if r.get("row_hash")
    ]
    return bulk_update_categories_by_row_hash(conn, updates, allow_clear=allow_clear)


def apply_income_categories_from_xlsx(
    conn,
    *,
    path: Path,
    allow_clear: bool = False,
) -> tuple[int, int]:
    """
    Updates the DB based on user edits in `raw_data/receitas.xlsx`.
    """
    rows = import_income_xlsx(path, source_hash="apply_income_xlsx", source_file=str(path))
    updates = [
        {
            "row_hash": r.get("row_hash"),
            "category": r.get("category"),
            "subcategory": r.get("subcategory"),
        }
        for r in rows
        if r.get("row_hash")
    ]
    return bulk_update_categories_by_row_hash(conn, updates, allow_clear=allow_clear)


def sync_credit_card_from_excel(
    conn,
    *,
    path: Path,
    card_owner_by_name: dict[str, str] | None = None,
) -> SyncResult:
    """
    Syncs the DB from the "Cartão" sheet in the unified Excel (Excel edits as source of
    truth for user fields).

    - Updates user-editable fields only (description, category, subcategory, person,
      notes) — never overwrites amounts or dates that came from a CSV import.
        - Soft-hides CC rows missing from the sheet (hidden_in_excel = 1), so deleting
            a line in Excel reflects in DB totals while preserving history.
    """
    file_hash = sha256_file(path)
    parsed = read_credit_card_master_xlsx(
        path,
        source_hash=file_hash,
        source_file=str(path),
        card_owner_by_name=card_owner_by_name,
    )
    return sync_transactions_by_row_hash(
        conn,
        payment_method="credit_card",
        rows=parsed.rows,
        delete_missing=True,
        user_fields_only=True,
    )


def sync_debit_from_unified_excel(conn, *, path: Path) -> SyncResult:
    """
    Syncs DB from the unified Excel (aba "Débitos").

    This sheet may include a `Hash (oculto)` column; if missing/empty we compute a deterministic hash.
    """
    if not _workbook_has_sheet(path, "Débitos"):
        return SyncResult(inserted=0, updated=0, deleted=0)
    file_hash = sha256_file(path)
    created = now_iso()
    raw = read_debitos_sheet(path)
    rows: list[dict] = []
    for r in raw:
        dt = r.get("date")
        description = str(r.get("description") or "").strip()
        amount_raw = r.get("amount")
        if dt is None or not description or amount_raw is None:
            continue
        amount = -abs(float(amount_raw))
        cash_dt = month_add(dt, 1)
        person = "Aline" if bool(r.get("pago_por_aline")) else "Renan"
        reimbursable = 1 if bool(r.get("reimbursable")) else 0
        notes = str(r.get("notes") or "").strip() or None
        category = str(r.get("category") or "").strip() or None
        subcategory = str(r.get("subcategory") or "").strip() or None

        row_hash = str(r.get("row_hash") or "").strip()
        if not row_hash:
            row_hash = sha256_text(
                "|".join(
                    [
                        "debit_unified",
                        dt.isoformat(),
                        f"{amount:.2f}",
                        description,
                    ]
                )
            )
        rows.append(
            {
                "origin_id": str(r.get("origin_id") or "").strip() or None,
                "row_hash": row_hash,
                "txn_date": dt.isoformat(),
                "cash_date": cash_dt.isoformat(),
                "amount": float(amount),
                "description": description,
                "group_name": None,
                "category": category,
                "subcategory": subcategory,
                "payment_method": "debit",
                "account": None,
                "source": "excel_unified_debit",
                "statement_closing_date": None,
                "statement_due_date": None,
                "person": person,
                "reimbursable": reimbursable,
                "reference": None,
                "notes": notes,
                "source_file": str(path),
                "source_hash": file_hash,
                "external_id": None,
                "created_at": created,
                "updated_at": created,
            }
        )
    return sync_transactions_by_row_hash(
        conn, payment_method="debit", rows=rows, delete_missing=True, user_fields_only=False
    )


def sync_income_from_unified_excel(conn, *, path: Path) -> SyncResult:
    """
    Syncs DB from the unified Excel (aba "Receitas").
    """
    if not _workbook_has_sheet(path, "Receitas"):
        return SyncResult(inserted=0, updated=0, deleted=0)
    file_hash = sha256_file(path)
    created = now_iso()
    raw = read_receitas_sheet(path)
    rows: list[dict] = []
    for r in raw:
        dt = r.get("date")
        description = str(r.get("description") or "").strip()
        amount_raw = r.get("amount")
        if dt is None or not description or amount_raw is None:
            continue
        amount = abs(float(amount_raw))
        cash_dt = month_add(dt, 1)
        person_raw = str(r.get("person") or "").strip()
        if person_raw:
            person = person_raw
        else:
            person = "Aline" if bool(r.get("recebido_por_aline")) else "Renan"
        notes = str(r.get("notes") or "").strip() or None
        category = str(r.get("category") or "").strip() or None

        row_hash = str(r.get("row_hash") or "").strip()
        if not row_hash:
            row_hash = sha256_text(
                "|".join(
                    [
                        "income_unified",
                        dt.isoformat(),
                        f"{amount:.2f}",
                        description,
                    ]
                )
            )
        rows.append(
            {
                "origin_id": str(r.get("origin_id") or "").strip() or None,
                "row_hash": row_hash,
                "txn_date": dt.isoformat(),
                "cash_date": cash_dt.isoformat(),
                "amount": float(amount),
                "description": description,
                "group_name": None,
                "category": category,
                "subcategory": None,
                "payment_method": "income",
                "account": None,
                "source": "excel_unified_income",
                "statement_closing_date": None,
                "statement_due_date": None,
                "person": person,
                "reimbursable": 0,
                "reference": None,
                "notes": notes,
                "source_file": str(path),
                "source_hash": file_hash,
                "external_id": None,
                "created_at": created,
                "updated_at": created,
            }
        )
    return sync_transactions_by_row_hash(
        conn, payment_method="income", rows=rows, delete_missing=True, user_fields_only=False
    )


def sync_household_from_unified_excel(
    conn,
    *,
    path: Path,
    tipo_to_subcategory: dict[str, str] | None = None,
) -> SyncResult:
    """
    Syncs DB from the unified Excel (aba "Contas Casa") into payment_method="household".

    Cash-date follows "Data Pagamento" (so the month filter matches the month the bill was paid),
    while "Mês Referência" is stored in `reference`.
    """
    if not _workbook_has_sheet(path, "Contas Casa"):
        return SyncResult(inserted=0, updated=0, deleted=0)
    file_hash = sha256_file(path)
    created = now_iso()
    raw = read_contas_casa_sheet(path)
    rows: list[dict] = []

    for r in raw:
        ref_month = str(r.get("reference_month") or "").strip() or None
        category = str(r.get("category") or "").strip()
        subcategory = str(r.get("subcategory") or "").strip() or None
        description = str(r.get("description") or "").strip()
        amount_raw = r.get("amount")
        if not category or amount_raw is None:
            continue

        payment_date = r.get("payment_date")
        if payment_date:
            cash_date = payment_date.isoformat()
            txn_date = cash_date
        elif ref_month:
            # Fallback: if payment_date not filled, pin to the reference month.
            cash_date = f"{ref_month}-01"
            txn_date = cash_date
        else:
            # Can't place it in any month without payment_date or reference month.
            continue

        amount = -abs(float(amount_raw))
        paid_by = str(r.get("paid_by") or "").strip() or None
        notes = str(r.get("notes") or "").strip() or None
        reference = ref_month or (payment_date.strftime("%Y-%m") if payment_date else None)

        row_hash = str(r.get("row_hash") or "").strip()
        if not row_hash:
            row_hash = sha256_text(
                "|".join(
                    [
                        "household_unified",
                        cash_date,
                        f"{amount:.2f}",
                        description,
                    ]
                )
            )

        rows.append(
            {
                "origin_id": str(r.get("origin_id") or "").strip() or None,
                "row_hash": row_hash,
                "txn_date": txn_date,
                "cash_date": cash_date,
                "amount": float(amount),
                "description": description or category,
                "group_name": None,
                "category": category,
                "subcategory": subcategory,
                "payment_method": "household",
                "account": None,
                "source": "excel_unified_household",
                "statement_closing_date": None,
                "statement_due_date": None,
                "person": paid_by,
                "reimbursable": 0,
                "reference": reference,
                "notes": notes,
                "source_file": str(path),
                "source_hash": file_hash,
                "external_id": None,
                "created_at": created,
                "updated_at": created,
            }
        )
    return sync_transactions_by_row_hash(
        conn, payment_method="household", rows=rows, delete_missing=True, user_fields_only=False
    )


def sync_unified_from_excel(
    conn,
    *,
    path: Path,
    card_owner_by_name: dict[str, str] | None = None,
) -> UnifiedSyncResult:
    return UnifiedSyncResult(
        credit_card=sync_credit_card_from_excel(conn, path=path, card_owner_by_name=card_owner_by_name),
        debit=sync_debit_from_unified_excel(conn, path=path),
        income=sync_income_from_unified_excel(conn, path=path),
        household=sync_household_from_unified_excel(conn, path=path),
    )

"""
Template unificado: um único arquivo Excel com abas para:
- Cartão de Crédito
- Débitos
- Receitas
- Contas da Casa (luz, água, babá, aluguel - pagas no início do mês)
"""
from __future__ import annotations

import io
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


def _uniq(xs: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in xs:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _excel_safe_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text or not re.match(r"^[a-z_]", text):
        text = f"c_{text}"
    return f"cat_{text[:50]}"


def _expense_categories(expense_tree: dict[str, Any]) -> tuple[list[str], dict[str, list[str]], dict[str, str]]:
    categories: list[str] = []
    sub_map: dict[str, list[str]] = {}
    safe_map: dict[str, str] = {}

    for category, subs in expense_tree.items():
        cat = str(category).strip()
        if not cat:
            continue
        categories.append(cat)
        safe_map[cat] = _excel_safe_name(cat)
        if isinstance(subs, list):
            sub_map[cat] = [str(x).strip() for x in subs if str(x).strip()]
        else:
            sub_map[cat] = []

    categories = _uniq(categories)
    for cat in categories:
        sub_map[cat] = _uniq(sub_map.get(cat, []))

    return categories, sub_map, safe_map


def _income_categories(income_tree: dict[str, Any]) -> list[str]:
    cats: list[str] = []
    for _, node in income_tree.items():
        if isinstance(node, list):
            cats.extend([str(x).strip() for x in node if str(x).strip()])
    return _uniq(cats)


def _add_defined_name(wb, dn) -> None:
    try:
        wb.defined_names.append(dn)
        return
    except Exception:
        pass

    add = getattr(wb.defined_names, "add", None)
    if callable(add):
        try:
            add(dn)
            return
        except TypeError:
            try:
                name = getattr(dn, "name", None)
                value = getattr(dn, "attr_text", None) or getattr(dn, "text", None)
                if name and value:
                    add(name, value)
                    return
            except Exception:
                pass
        except Exception:
            pass

    try:
        wb.defined_names[dn.name] = dn
    except Exception:
        return


def build_unified_template_bytes(
    *,
    expense_categories_tree: dict[str, Any],
    income_categories_tree: dict[str, Any],
    cards_list: list[str],
) -> bytes:
    """
    Cria um único arquivo Excel com 4 abas:
    - Cartão de Crédito
    - Débitos
    - Receitas
    - Contas da Casa
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # Create sheets
    ws_cartao = wb.create_sheet("Cartão de Crédito")
    ws_debito = wb.create_sheet("Débitos")
    ws_receita = wb.create_sheet("Receitas")
    ws_contas = wb.create_sheet("Contas da Casa")
    ws_lists = wb.create_sheet("Listas")
    
    # Remove default sheet
    wb.remove(default_sheet)
    
    # --- Setup Lists sheet ---
    categories, sub_map, safe_map = _expense_categories(expense_categories_tree)
    income_cats = _income_categories(income_categories_tree)
    
    # Column A: Categories
    for i, cat in enumerate(categories, start=1):
        ws_lists.cell(row=i, column=1, value=cat)
    
    # Column B: Safe names for categories
    for i, cat in enumerate(categories, start=1):
        ws_lists.cell(row=i, column=2, value=safe_map[cat])
    
    # Subcategories in columns starting at C
    base_col = 3
    for idx, cat in enumerate(categories):
        col = base_col + idx
        key = safe_map[cat]
        ws_lists.cell(row=1, column=col, value=key)
        subs = sub_map.get(cat, [])
        if not subs:
            ws_lists.cell(row=2, column=col, value="")
            subs = [""]
        for r_i, sub in enumerate(subs, start=2):
            ws_lists.cell(row=r_i, column=col, value=sub)
        
        # Define named range for subcategories
        col_letter = get_column_letter(col)
        start_ref = f"${col_letter}$2"
        end_ref = f"${col_letter}${1 + len(subs)}"
        dn = DefinedName(key, attr_text=f"Listas!{start_ref}:{end_ref}")
        _add_defined_name(wb, dn)
    
    # Define named range for categories
    if categories:
        dn = DefinedName("Categorias", attr_text=f"Listas!$A$1:$A${len(categories)}")
        _add_defined_name(wb, dn)
    
    # Person list in a dedicated column
    pessoas = ["Renan", "Aline", "Família"]
    pessoa_col = base_col + len(categories) + 2
    ws_lists.cell(row=1, column=pessoa_col, value="Pessoas")
    for i, p in enumerate(pessoas, start=2):
        ws_lists.cell(row=i, column=pessoa_col, value=p)
    dn = DefinedName("Pessoas", attr_text=f"Listas!${get_column_letter(pessoa_col)}$2:${get_column_letter(pessoa_col)}${1 + len(pessoas)}")
    _add_defined_name(wb, dn)
    
    # Cards list
    cartao_col = pessoa_col + 1
    ws_lists.cell(row=1, column=cartao_col, value="Cartões")
    for i, card in enumerate(cards_list, start=2):
        ws_lists.cell(row=i, column=cartao_col, value=card)
    if cards_list:
        dn = DefinedName("Cartoes", attr_text=f"Listas!${get_column_letter(cartao_col)}$2:${get_column_letter(cartao_col)}${1 + len(cards_list)}")
        _add_defined_name(wb, dn)
    
    # Income categories
    income_col = cartao_col + 1
    ws_lists.cell(row=1, column=income_col, value="CategoriasReceita")
    for i, cat in enumerate(income_cats, start=2):
        ws_lists.cell(row=i, column=income_col, value=cat)
    if income_cats:
        dn = DefinedName("CategoriasReceita", attr_text=f"Listas!${get_column_letter(income_col)}$2:${get_column_letter(income_col)}${1 + len(income_cats)}")
        _add_defined_name(wb, dn)
    
    # Contas da casa list
    contas_casa = ["Aluguel", "Energia", "Água", "Babá", "Internet", "Gás", "Conta de Celular"]
    contas_col = income_col + 1
    ws_lists.cell(row=1, column=contas_col, value="ContasCasa")
    for i, conta in enumerate(contas_casa, start=2):
        ws_lists.cell(row=i, column=contas_col, value=conta)
    dn = DefinedName("ContasCasa", attr_text=f"Listas!${get_column_letter(contas_col)}$2:${get_column_letter(contas_col)}${1 + len(contas_casa)}")
    _add_defined_name(wb, dn)
    
    # Status list
    status_col = contas_col + 1
    ws_lists.cell(row=1, column=status_col, value="Status")
    for i, s in enumerate(["Pago", "Pendente"], start=2):
        ws_lists.cell(row=i, column=status_col, value=s)
    dn = DefinedName("StatusPago", attr_text=f"Listas!${get_column_letter(status_col)}$2:${get_column_letter(status_col)}$3")
    _add_defined_name(wb, dn)
    
    ws_lists.sheet_state = "hidden"
    
    # --- Cartão de Crédito sheet ---
    _setup_credit_card_sheet(wb, ws_cartao, categories, safe_map)
    
    # --- Débitos sheet ---
    _setup_debit_sheet(wb, ws_debito, categories, safe_map)
    
    # --- Receitas sheet ---
    _setup_income_sheet(wb, ws_receita)
    
    # --- Contas da Casa sheet ---
    _setup_household_sheet(wb, ws_contas, categories, safe_map)
    
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _setup_credit_card_sheet(wb, ws, categories: list[str], safe_map: dict[str, str]):
    """Configura a aba de Cartão de Crédito"""
    headers = [
        "Data da compra",
        "Data do vencimento",
        "Categoria",
        "Subcategoria",
        "Cartão de crédito",
        "Descrição",
        "Valor",
        "Status",
        "Pessoa",
        "__chave_categoria (oculto)",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    
    # Data validations
    dv_cat = DataValidation(type="list", formula1="=Categorias", allow_blank=True)
    dv_sub = DataValidation(type="list", formula1="=INDIRECT($J2)", allow_blank=True)
    dv_card = DataValidation(type="list", formula1="=Cartoes", allow_blank=True)
    dv_status = DataValidation(type="list", formula1="=StatusPago", allow_blank=True)
    dv_person = DataValidation(type="list", formula1="=Pessoas", allow_blank=True)
    
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_sub)
    ws.add_data_validation(dv_card)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_person)
    
    start_row, end_row = 2, 5000
    dv_cat.add(f"C{start_row}:C{end_row}")
    dv_sub.add(f"D{start_row}:D{end_row}")
    dv_card.add(f"E{start_row}:E{end_row}")
    dv_status.add(f"H{start_row}:H{end_row}")
    dv_person.add(f"I{start_row}:I{end_row}")
    
    # Default values and formulas
    lookup_end = max(1, len(categories))
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=8, value="Pago")
        ws.cell(row=r, column=9, value="Renan")
        ws.cell(row=r, column=10, value=f'=IFERROR(VLOOKUP($C{r},Listas!$A$1:$B${lookup_end},2,FALSE),"")')
    
    ws.column_dimensions["J"].hidden = True
    
    widths = [14, 14, 28, 28, 22, 40, 14, 12, 12, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _setup_debit_sheet(wb, ws, categories: list[str], safe_map: dict[str, str]):
    """Configura a aba de Débitos"""
    headers = [
        "Data",
        "Categoria",
        "Subcategoria",
        "Descrição",
        "Valor",
        "Status",
        "Pago por",
        "__chave_categoria (oculto)",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    
    dv_cat = DataValidation(type="list", formula1="=Categorias", allow_blank=True)
    dv_sub = DataValidation(type="list", formula1="=INDIRECT($H2)", allow_blank=True)
    dv_status = DataValidation(type="list", formula1="=StatusPago", allow_blank=True)
    dv_person = DataValidation(type="list", formula1="=Pessoas", allow_blank=True)
    
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_sub)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_person)
    
    start_row, end_row = 2, 5000
    dv_cat.add(f"B{start_row}:B{end_row}")
    dv_sub.add(f"C{start_row}:C{end_row}")
    dv_status.add(f"F{start_row}:F{end_row}")
    dv_person.add(f"G{start_row}:G{end_row}")
    
    lookup_end = max(1, len(categories))
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=6, value="Pago")
        ws.cell(row=r, column=7, value="Aline")
        ws.cell(row=r, column=8, value=f'=IFERROR(VLOOKUP($B{r},Listas!$A$1:$B${lookup_end},2,FALSE),"")')
    
    ws.column_dimensions["H"].hidden = True
    
    widths = [14, 28, 28, 40, 14, 12, 12, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _setup_income_sheet(wb, ws):
    """Configura a aba de Receitas"""
    headers = [
        "Data",
        "Descrição",
        "Valor",
        "Categoria",
        "Pessoa",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    
    dv_cat = DataValidation(type="list", formula1="=CategoriasReceita", allow_blank=True)
    dv_person = DataValidation(type="list", formula1="=Pessoas", allow_blank=True)
    
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_person)
    
    start_row, end_row = 2, 5000
    dv_cat.add(f"D{start_row}:D{end_row}")
    dv_person.add(f"E{start_row}:E{end_row}")
    
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=5, value="Renan")
    
    widths = [14, 40, 14, 28, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _setup_household_sheet(wb, ws, categories: list[str], safe_map: dict[str, str]):
    """
    Configura a aba de Contas da Casa.
    Essas contas são divididas no início do mês corrente (não no acerto do mês anterior).
    """
    headers = [
        "Mês de Referência",
        "Categoria",
        "Subcategoria",
        "Descrição",
        "Valor",
        "Data de Pagamento",
        "Pago por",
        "Status",
        "__chave_categoria (oculto)",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    
    # Add note explaining the purpose
    ws.cell(row=1, column=11, value="← Contas divididas no início do mês corrente")
    
    dv_cat = DataValidation(type="list", formula1="=Categorias", allow_blank=True)
    dv_sub = DataValidation(type="list", formula1="=INDIRECT($I2)", allow_blank=True)
    dv_person = DataValidation(type="list", formula1="=Pessoas", allow_blank=True)
    dv_status = DataValidation(type="list", formula1="=StatusPago", allow_blank=True)
    
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_sub)
    ws.add_data_validation(dv_person)
    ws.add_data_validation(dv_status)
    
    start_row, end_row = 2, 500
    dv_cat.add(f"B{start_row}:B{end_row}")
    dv_sub.add(f"C{start_row}:C{end_row}")
    dv_person.add(f"G{start_row}:G{end_row}")
    dv_status.add(f"H{start_row}:H{end_row}")
    
    lookup_end = max(1, len(categories))
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=7, value="Aline")
        ws.cell(row=r, column=8, value="Pago")
        ws.cell(row=r, column=9, value=f'=IFERROR(VLOOKUP($B{r},Listas!$A$1:$B${lookup_end},2,FALSE),"")')
    
    ws.column_dimensions["I"].hidden = True
    
    widths = [18, 28, 28, 40, 14, 18, 12, 12, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def ensure_unified_template(
    base_dir: Path,
    *,
    expense_categories_tree: dict[str, Any],
    income_categories_tree: dict[str, Any],
    cards_list: list[str],
    force: bool = False,
) -> Path:
    """
    Cria o template unificado em raw_data/lancamentos.xlsx
    """
    raw_dir = base_dir / "raw_data"
    raw_dir.mkdir(parents=True, exist_ok=True)
    path = raw_dir / "lancamentos.xlsx"
    
    if force or not path.exists():
        content = build_unified_template_bytes(
            expense_categories_tree=expense_categories_tree,
            income_categories_tree=income_categories_tree,
            cards_list=cards_list,
        )
        path.write_bytes(content)
    
    return path

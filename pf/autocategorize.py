from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from pf.utils import normalize_str


@dataclass(frozen=True)
class Suggestion:
    category: str
    subcategory: str


def _pick_default_for_owner(owner: str | None) -> Suggestion:
    owner_n = normalize_str(owner)
    if owner_n == "aline":
        return Suggestion(category="Gastos Aline", subcategory="Geral")
    if owner_n == "familia":
        return Suggestion(category="Gastos Variáveis", subcategory="Residual")
    return Suggestion(category="Gastos Renan", subcategory="Variados Renan (residual)")


def _ensure_valid(s: Suggestion, expense_categories_tree: dict[str, list[str]]) -> Suggestion:
    if s.category not in expense_categories_tree:
        return Suggestion(category="Gastos Variáveis", subcategory="Residual")
    subs = expense_categories_tree.get(s.category) or []
    if not subs:
        return s
    if s.subcategory in subs:
        return s
    # fallback: keep a deterministic option
    return Suggestion(category=s.category, subcategory=subs[0])


def suggest_credit_card_category_subcategory(
    *,
    description: str,
    account: str | None,
    expense_categories_tree: dict[str, list[str]],
    card_owner_by_name: dict[str, str] | None = None,
) -> Suggestion:
    desc = normalize_str(description)
    owner = None
    if card_owner_by_name and account:
        owner = card_owner_by_name.get(str(account).strip())

    # Travel / lodging
    if any(k in desc for k in ("airbnb", "hotel", "hosped")):
        return _ensure_valid(Suggestion("Lazer", "Hospedagens"), expense_categories_tree)
    if any(k in desc for k in ("gol linhas", "smiles")):
        return _ensure_valid(Suggestion("Lazer", "Passagens"), expense_categories_tree)

    # Fuel / car
    if any(k in desc for k in ("uber", "99app", "99 pop", "99pop")):
        return _ensure_valid(Suggestion("Transporte", "Aplicativo"), expense_categories_tree)
    if "nutag" in desc:
        return _ensure_valid(Suggestion("Transporte", "Pedágio"), expense_categories_tree)
    if any(k in desc for k in ("posto", "shellbox", "combust")):
        return _ensure_valid(Suggestion("Transporte", "Combustível"), expense_categories_tree)
    if "porto seguro" in desc:
        return _ensure_valid(Suggestion("Transporte", "Seguro"), expense_categories_tree)
    if any(k in desc for k in ("motonet", "mecan", "oficina", "auto posto")):
        return _ensure_valid(Suggestion("Transporte", "Mecânico"), expense_categories_tree)

    # Groceries / food at home
    if any(
        k in desc
        for k in (
            "supermerc",
            "tenda atacado",
            "atacado",
            "mercado",
            "martins",
            "carne",
            "espigraos",
            "ponto do ovo",
            "padaria",
            "cereal",
        )
    ):
        return _ensure_valid(Suggestion("Alimentação", "Mercado"), expense_categories_tree)

    # Pharmacy / health
    if any(k in desc for k in ("raia", "drogasil", "drogaria", "rd saude", "pharm")):
        return _ensure_valid(Suggestion("Saúde", "Farmácia Geral"), expense_categories_tree)

    # Restaurants / coffee
    if "ifd*" in desc and not any(k in desc for k in ("drogaria", "drogasil", "raia", "rd saude")):
        return _ensure_valid(Suggestion("Lazer", "Restaurantes"), expense_categories_tree)
    if any(
        k in desc
        for k in (
            "pizz",
            "restaur",
            "adega",
            "culinaria",
            "gelato",
            "sorvet",
            "alecrim e tomilho",
            "flor de maria",
        )
    ):
        sub = "Cafés" if any(k in desc for k in ("gelato", "sorvet", "cafe", "caf")) else "Restaurantes"
        return _ensure_valid(Suggestion("Lazer", sub), expense_categories_tree)

    # Mobile plan
    if any(k in desc for k in ("nucel", "plano nucel", "45gb", "celular")):
        return _ensure_valid(Suggestion("Habitação", "Conta de Celular"), expense_categories_tree)

    # E-commerce / household items
    if any(k in desc for k in ("casasbahia", "casas bahia")):
        return _ensure_valid(Suggestion("Gastos Variáveis", "Eletrodomésticos"), expense_categories_tree)
    if any(k in desc for k in ("casa e cia", "casa & cia", "casa e cia itu")):
        return _ensure_valid(Suggestion("Gastos Variáveis", "Casa – Itens & Utilidades"), expense_categories_tree)

    # Baby/kid items
    if any(k in desc for k in ("anababy", "baby", "cresci e perdi", "fralda")):
        sub = "Fraldas" if "fralda" in desc else "Roupinhas"
        return _ensure_valid(Suggestion("Despesas Dudu", sub), expense_categories_tree)

    # Clothes (fallback by owner)
    if any(k in desc for k in ("renner", "riachuelo", "cea", "comercio de roupas", "comercio", "roup")):
        owner_n = normalize_str(owner)
        if owner_n == "aline":
            return _ensure_valid(Suggestion("Gastos Aline", "Vestuário"), expense_categories_tree)
        if owner_n == "renan":
            return _ensure_valid(Suggestion("Gastos Renan", "Roupas"), expense_categories_tree)
        return _ensure_valid(Suggestion("Gastos Variáveis", "Residual"), expense_categories_tree)

    # Subscriptions / services
    if any(k in desc for k in ("github", "investidor10", "assin")):
        return _ensure_valid(Suggestion("Gastos Renan", "Assinaturas Pessoais"), expense_categories_tree)

    return _ensure_valid(_pick_default_for_owner(owner), expense_categories_tree)


def autofill_unified_credit_card_sheet(
    path: Path,
    *,
    expense_categories_tree: dict[str, list[str]],
    card_owner_by_name: dict[str, str] | None = None,
    override: bool = False,
) -> int:
    """
    Preenche `Categoria` e `Subcategoria` na aba "Cartão" do Excel unificado, apenas onde estiver vazio.
    """
    wb = load_workbook(path)
    if "Cartão" not in wb.sheetnames:
        raise ValueError('Aba "Cartão" não encontrada no Excel.')
    ws = wb["Cartão"]

    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        key = normalize_str(v)
        if key:
            headers[key] = c

    cat_col = headers.get(normalize_str("Categoria"))
    sub_col = headers.get(normalize_str("Subcategoria"))
    desc_col = headers.get(normalize_str("Descrição")) or headers.get(normalize_str("Descricao"))
    acct_col = headers.get(normalize_str("Cartão de crédito")) or headers.get(normalize_str("Cartao de credito"))

    if not cat_col or not sub_col or not desc_col:
        raise ValueError("Colunas esperadas não encontradas na aba Cartão (Categoria/Subcategoria/Descrição).")

    rows_updated = 0
    for r in range(2, ws.max_row + 1):
        desc = str(ws.cell(row=r, column=desc_col).value or "").strip()
        if not desc:
            continue
        cat = str(ws.cell(row=r, column=cat_col).value or "").strip()
        sub = str(ws.cell(row=r, column=sub_col).value or "").strip()
        if not override and cat and sub:
            continue

        account = None
        if acct_col:
            account = str(ws.cell(row=r, column=acct_col).value or "").strip() or None

        sug = suggest_credit_card_category_subcategory(
            description=desc,
            account=account,
            expense_categories_tree=expense_categories_tree,
            card_owner_by_name=card_owner_by_name,
        )

        wrote_any = False
        if override:
            if cat != sug.category:
                ws.cell(row=r, column=cat_col).value = sug.category
                wrote_any = True
            if sub != sug.subcategory:
                ws.cell(row=r, column=sub_col).value = sug.subcategory
                wrote_any = True
        else:
            if not cat:
                ws.cell(row=r, column=cat_col).value = sug.category
                wrote_any = True
            if not sub:
                ws.cell(row=r, column=sub_col).value = sug.subcategory
                wrote_any = True

        if wrote_any:
            rows_updated += 1

    if rows_updated:
        wb.save(path)
    return rows_updated

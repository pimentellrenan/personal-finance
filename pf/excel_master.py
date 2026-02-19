from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook

from pf.utils import normalize_str, parse_date


def _find_col(headers: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for cand in candidates:
        key = normalize_str(cand)
        if key in headers:
            return headers[key]
    for cand in candidates:
        key = normalize_str(cand)
        for h, idx in headers.items():
            if key and key in h:
                return idx
    return None


def _ensure_credit_card_master(path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cartão"
    ws.append(
        [
            "Hash (oculto)",
            "Data da compra",
            "Data do vencimento",
            "Categoria",
            "Subcategoria",
            "Cartão de crédito",
            "Descrição",
            "Valor (R$)",
            "Status (cartao pago ou nao)",
        ]
    )
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].hidden = True
    wb.save(path)


def append_credit_card_rows_to_master_xlsx(
    path: Path,
    *,
    rows: Iterable[dict[str, Any]],
) -> int:
    """
    Appends credit-card rows into the Excel master workbook (`templates/cartao_credito.xlsx`)
    without overwriting existing user edits.

    Rows are deduped by `row_hash`.
    Returns the number of appended rows.
    """
    _ensure_credit_card_master(path)
    wb = load_workbook(path)
    sheet_name = "Cartão" if "Cartão" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        key = normalize_str(v)
        if key:
            headers[key] = c

    hash_col = _find_col(headers, ("hash (oculto)", "hash", "row_hash", "id")) or 1
    txn_col = _find_col(headers, ("data da compra", "data compra", "txn_date", "data")) or 2
    due_col = _find_col(headers, ("data do vencimento", "vencimento", "statement_due_date")) or 3
    cat_col = _find_col(headers, ("categoria", "category")) or 4
    sub_col = _find_col(headers, ("subcategoria", "sub-category", "sub_category", "subcategory")) or 5
    card_col = _find_col(headers, ("cartao de credito", "cartão de crédito", "cartao", "cartão", "account")) or 6
    desc_col = _find_col(headers, ("descricao", "descrição", "description")) or 7
    amount_col = _find_col(headers, ("valor", "valor (r$)", "valor (em r$)", "amount")) or 8
    status_col = _find_col(headers, ("status (cartao pago ou nao)", "status")) or 9

    existing: set[str] = set()
    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        rh = str(ws.cell(row=r, column=hash_col).value or "").strip()
        if rh:
            existing.add(rh)
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in (txn_col, due_col, desc_col, amount_col)):
            last_data_row = r

    appended = 0
    today = date.today()
    next_row = last_data_row + 1

    for row in rows:
        rh = str(row.get("row_hash") or "").strip()
        if not rh or rh in existing:
            continue

        txn_date = parse_date(row.get("txn_date"))
        due_date = parse_date(row.get("statement_due_date") or row.get("cash_date"))
        description = str(row.get("description") or "").strip()
        account = str(row.get("account") or "").strip()
        category = str(row.get("category") or "").strip()
        subcategory = str(row.get("subcategory") or "").strip()

        amt = row.get("amount")
        try:
            amount_v = float(amt) if amt is not None else None
        except Exception:  # noqa: BLE001
            amount_v = None

        if txn_date is None or due_date is None or amount_v is None or not description or not account:
            continue

        ws.cell(row=next_row, column=hash_col).value = rh
        ws.cell(row=next_row, column=txn_col).value = txn_date
        ws.cell(row=next_row, column=due_col).value = due_date
        ws.cell(row=next_row, column=cat_col).value = category
        ws.cell(row=next_row, column=sub_col).value = subcategory
        ws.cell(row=next_row, column=card_col).value = account
        ws.cell(row=next_row, column=desc_col).value = description
        ws.cell(row=next_row, column=amount_col).value = abs(amount_v)

        status = "Pago" if due_date <= today else "Em aberto"
        ws.cell(row=next_row, column=status_col).value = status

        existing.add(rh)
        appended += 1
        next_row += 1

    if appended:
        wb.save(path)
    return appended

from __future__ import annotations

import io
import re
import unicodedata
from pathlib import Path
from typing import Any


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(
            "Dependência ausente: openpyxl. Instale com `pip install -r requirements.txt`."
        ) from e


def _uniq(xs: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in xs:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _excel_safe_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text or not re.match(r"^[a-z_]", text):
        text = f"c_{text}"
    return f"cat_{text[:50]}"


def _expense_categories(expense_tree: dict[str, Any]) -> tuple[list[str], dict[str, list[str]], dict[str, str]]:
    categories: list[str] = []
    sub_map: dict[str, list[str]] = {}
    safe_map: dict[str, str] = {}

    for category, subs in expense_tree.items():
        cat = str(category).strip()
        if not cat:
            continue
        categories.append(cat)
        safe_map[cat] = _excel_safe_name(cat)
        if isinstance(subs, list):
            sub_map[cat] = [str(x).strip() for x in subs if str(x).strip()]
        else:
            sub_map[cat] = []

    categories = _uniq(categories)
    for cat in categories:
        sub_map[cat] = _uniq(sub_map.get(cat, []))

    return categories, sub_map, safe_map


def _income_categories(income_tree: dict[str, Any]) -> list[str]:
    # Supports:
    # - {"Receitas": [..]}
    # - {"Receitas": {"...": [...]}} (ignored)
    cats: list[str] = []
    for _, node in income_tree.items():
        if isinstance(node, list):
            cats.extend([str(x).strip() for x in node if str(x).strip()])
    return _uniq(cats)


def _add_defined_name(wb, dn) -> None:
    """
    openpyxl compatibility: `Workbook.defined_names` changed across versions
    (list-like vs dict-like).
    """
    # Older: DefinedNameList (list-like)
    try:
        wb.defined_names.append(dn)
        return
    except Exception:  # noqa: BLE001
        pass

    # Newer: DefinedNameDict (has .add and/or mapping interface)
    add = getattr(wb.defined_names, "add", None)
    if callable(add):
        try:
            add(dn)
            return
        except TypeError:
            # Some versions use add(name, value)
            try:
                name = getattr(dn, "name", None)
                value = getattr(dn, "attr_text", None) or getattr(dn, "text", None)
                if name and value:
                    add(name, value)
                    return
            except Exception:  # noqa: BLE001
                pass
        except Exception:  # noqa: BLE001
            pass

    try:
        wb.defined_names[dn.name] = dn
    except Exception:  # noqa: BLE001
        return


def build_debit_template_bytes(*, expense_categories_tree: dict[str, Any]) -> bytes:
    _require_openpyxl()
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws_lists = wb.create_sheet("Listas")

    headers = [
        "Data da compra",
        "Data do vencimento",
        "Categoria",
        "Subcategoria",
        "Cartão de crédito",
        "Descrição",
        "Valor",
        "Status (cartao pago ou nao)",
        "Pessoa",
    ]
    # Hidden helper column to implement dependent dropdowns in Excel (Categoria -> Subcategoria).
    headers.append("__chave_categoria (oculto)")
    ws.append(headers)
    ws.freeze_panes = "A2"

    categories, sub_map, safe_map = _expense_categories(expense_categories_tree)
    for i, cat in enumerate(categories, start=1):
        ws_lists.cell(row=i, column=1, value=cat)
        ws_lists.cell(row=i, column=2, value=safe_map[cat])

    # Put each subcategory list in its own column starting at C (3)
    base_col = 3
    max_rows = 2
    for idx, cat in enumerate(categories):
        col = base_col + idx
        key = safe_map[cat]
        ws_lists.cell(row=1, column=col, value=key)
        subs = sub_map.get(cat, [])
        if not subs:
            ws_lists.cell(row=2, column=col, value="")
            subs = [""]
        for r_i, sub in enumerate(subs, start=2):
            ws_lists.cell(row=r_i, column=col, value=sub)
        max_rows = max(max_rows, 1 + len(subs))

        # Define named range for this category key (subcategories list)
        col_letter = get_column_letter(col)
        start_ref = f"${col_letter}$2"
        end_ref = f"${col_letter}${1 + len(subs)}"
        dn = DefinedName(key, attr_text=f"Listas!{start_ref}:{end_ref}")
        _add_defined_name(wb, dn)

    # Define named range for categories list
    if categories:
        dn = DefinedName("Categorias", attr_text=f"Listas!$A$1:$A${len(categories)}")
        _add_defined_name(wb, dn)

    # Add person list (Renan, Aline, Família)
    pessoas = ["Renan", "Aline", "Família"]
    pessoa_col = max(base_col + len(categories) + 1, 50)  # Use column far to the right
    for i, p in enumerate(pessoas, start=1):
        ws_lists.cell(row=i, column=pessoa_col, value=p)
    
    # Define named range for person list
    dn = DefinedName("Pessoas", attr_text=f"Listas!${get_column_letter(pessoa_col)}$1:${get_column_letter(pessoa_col)}${len(pessoas)}")
    _add_defined_name(wb, dn)

    # Data validations
    dv_cat = DataValidation(type="list", formula1="=Categorias", allow_blank=True)
    dv_sub = DataValidation(type="list", formula1="=INDIRECT($J2)", allow_blank=True)
    dv_person = DataValidation(type="list", formula1="=Pessoas", allow_blank=True)

    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_sub)
    ws.add_data_validation(dv_person)

    # Apply to a generous range.
    start_row = 2
    end_row = 5000
    dv_cat.add(f"C{start_row}:C{end_row}")
    dv_sub.add(f"D{start_row}:D{end_row}")
    dv_person.add(f"I{start_row}:I{end_row}")

    # Default "Pago" for status and "Renan" for person + formulas for hidden key column.
    lookup_end = max(1, len(categories))
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=8, value="Pago")
        ws.cell(row=r, column=9, value="Renan")  # Default person
        ws.cell(
            row=r,
            column=10,
            value=f'=IFERROR(VLOOKUP($C{r},Listas!$A$1:$B${lookup_end},2,FALSE),"")',
        )

    ws.column_dimensions["J"].hidden = True

    # Some basic column widths
    widths = [14, 14, 28, 28, 22, 40, 16, 22, 15, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w

    ws_lists.sheet_state = "hidden"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_income_template_bytes(*, income_categories_tree: dict[str, Any]) -> bytes:
    _require_openpyxl()
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws_lists = wb.create_sheet("Listas")

    headers = [
        "Data (AAAA-MM-DD)",
        "Descrição",
        "Valor (positivo)",
        "Categoria",
        "Subcategoria",
        "Pessoa",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"

    categories = _income_categories(income_categories_tree)
    for i, c in enumerate(categories, start=1):
        ws_lists.cell(row=i, column=1, value=c)

    # Add person list (Renan, Aline, Família)
    pessoas = ["Renan", "Aline", "Família"]
    for i, p in enumerate(pessoas, start=1):
        ws_lists.cell(row=i, column=2, value=p)

    dv_cat = DataValidation(type="list", formula1=f"=Listas!$A$1:$A${max(1, len(categories))}", allow_blank=True)
    dv_person = DataValidation(type="list", formula1=f"=Listas!$B$1:$B${len(pessoas)}", allow_blank=True)

    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_person)

    start_row = 2
    end_row = 5000
    dv_cat.add(f"D{start_row}:D{end_row}")
    dv_person.add(f"F{start_row}:F{end_row}")
    
    # Default "Renan" for person
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=6, value="Renan")

    widths = [14, 40, 16, 28, 24, 15]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w

    ws_lists.sheet_state = "hidden"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def ensure_templates_on_disk(
    base_dir: Path,
    *,
    expense_categories_tree: dict[str, Any],
    income_categories_tree: dict[str, Any],
) -> dict[str, Path]:
    templates_dir = base_dir / "templates"
    templates_dir.mkdir(parents=True, exist_ok=True)
    debit_path = templates_dir / "debitos.xlsx"
    income_path = templates_dir / "receitas.xlsx"

    if not debit_path.exists():
        debit_path.write_bytes(build_debit_template_bytes(expense_categories_tree=expense_categories_tree))
    if not income_path.exists():
        income_path.write_bytes(build_income_template_bytes(income_categories_tree=income_categories_tree))

    return {"debit": debit_path, "income": income_path}


def ensure_input_templates_on_disk(
    base_dir: Path,
    *,
    expense_categories_tree: dict[str, Any],
    income_categories_tree: dict[str, Any],
) -> dict[str, Path]:
    """
    Creates the input templates inside `raw_data/` so the user can fill them and
    re-run the import without uploads.
    """
    raw_dir = base_dir / "raw_data"
    raw_dir.mkdir(parents=True, exist_ok=True)
    debit_path = raw_dir / "debitos.xlsx"
    income_path = raw_dir / "receitas.xlsx"

    if not debit_path.exists():
        debit_path.write_bytes(build_debit_template_bytes(expense_categories_tree=expense_categories_tree))
    if not income_path.exists():
        income_path.write_bytes(build_income_template_bytes(income_categories_tree=income_categories_tree))

    return {"debit": debit_path, "income": income_path}

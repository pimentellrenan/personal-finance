"""
Excel Unificado - Todas as transações em um único arquivo com múltiplas abas:
1. Cartão - Lançamentos de cartão de crédito
2. Débitos - Gastos pagos via débito/PIX/dinheiro
3. Receitas - Rendimentos e entradas
4. Contas Casa - Contas fixas da casa que entram no acerto do mês em que são pagas (com "Mês Referência")
"""
from __future__ import annotations

import io
import re
import unicodedata
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from pf.utils import normalize_str, parse_date, parse_brl_number


# ============================================================================
# Helpers
# ============================================================================

def _excel_safe_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text or not re.match(r"^[a-z_]", text):
        text = f"c_{text}"
    return f"cat_{text[:50]}"


def _uniq(xs: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in xs:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _expense_categories(expense_tree: dict[str, Any]) -> tuple[list[str], dict[str, list[str]], dict[str, str]]:
    categories: list[str] = []
    sub_map: dict[str, list[str]] = {}
    safe_map: dict[str, str] = {}

    for category, subs in expense_tree.items():
        cat = str(category).strip()
        if not cat:
            continue
        categories.append(cat)
        safe_map[cat] = _excel_safe_name(cat)
        if isinstance(subs, list):
            sub_map[cat] = [str(x).strip() for x in subs if str(x).strip()]
        else:
            sub_map[cat] = []

    categories = _uniq(categories)
    for cat in categories:
        sub_map[cat] = _uniq(sub_map.get(cat, []))

    return categories, sub_map, safe_map


def _income_categories(income_tree: dict[str, Any]) -> list[str]:
    cats: list[str] = []
    for _, node in income_tree.items():
        if isinstance(node, list):
            cats.extend([str(x).strip() for x in node if str(x).strip()])
    return _uniq(cats)


def _add_defined_name(wb, dn) -> None:
    try:
        wb.defined_names.append(dn)
        return
    except Exception:
        pass
    add = getattr(wb.defined_names, "add", None)
    if callable(add):
        try:
            add(dn)
            return
        except TypeError:
            try:
                name = getattr(dn, "name", None)
                value = getattr(dn, "attr_text", None) or getattr(dn, "text", None)
                if name and value:
                    add(name, value)
                    return
            except Exception:
                pass
    try:
        wb.defined_names[dn.name] = dn
    except Exception:
        return


def _find_col(headers: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for cand in candidates:
        key = normalize_str(cand)
        if key in headers:
            return headers[key]
    for cand in candidates:
        key = normalize_str(cand)
        for h, idx in headers.items():
            if key and key in h:
                return idx
    return None


_PT_MONTHS: dict[str, int] = {
    "jan": 1,
    "janeiro": 1,
    "fev": 2,
    "fevereiro": 2,
    "mar": 3,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "abr": 4,
    "mai": 5,
    "maio": 5,
    "jun": 6,
    "junho": 6,
    "jul": 7,
    "julho": 7,
    "ago": 8,
    "agosto": 8,
    "set": 9,
    "setembro": 9,
    "out": 10,
    "outubro": 10,
    "nov": 11,
    "novembro": 11,
    "dez": 12,
    "dezembro": 12,
}


def _parse_reference_month(value: Any) -> str | None:
    """
    Parse "Mês Referência" into "YYYY-MM".

    Accepts values like:
    - 2026-01 / 2026/01
    - 01/2026
    - jan/26, fev/2026 (PT-BR month names)
    - any date/datetime (uses its year/month)
    """
    if value is None:
        return None

    d = parse_date(value)
    if d:
        return f"{d.year}-{d.month:02d}"

    text = str(value).strip()
    if not text:
        return None

    s = normalize_str(text)

    # YYYY-MM or YYYY/MM
    m = re.match(r"^\s*(\d{4})\s*[-/]\s*(\d{1,2})\s*$", s)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        if 1 <= month <= 12:
            return f"{year}-{month:02d}"

    # MM/YYYY
    m = re.match(r"^\s*(\d{1,2})\s*[-/]\s*(\d{4})\s*$", s)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        if 1 <= month <= 12:
            return f"{year}-{month:02d}"

    # month_name/YY or month_name/YYYY (e.g., jan/26, fev/2026)
    m = re.match(r"^\s*([a-zç]+)\s*[-/]\s*(\d{2,4})\s*$", s)
    if m:
        mon_raw = m.group(1).strip().lower()
        year_raw = m.group(2).strip()
        month = _PT_MONTHS.get(mon_raw)
        if month is None and len(mon_raw) >= 3:
            month = _PT_MONTHS.get(mon_raw[:3])
        if month is None:
            return None
        year = int(year_raw)
        if len(year_raw) == 2:
            year += 2000
        return f"{year}-{month:02d}"

    return None


# ============================================================================
# Contas da Casa - lista fixa
# ============================================================================

CONTAS_CASA = [
    "Luz",
    "Água", 
    "Babá",
    "Aluguel",
    "Internet",
    "Gás",
    "Condomínio",
    "IPTU",
    "Seguro Casa",
    "Outros",
]


# ============================================================================
# Build Unified Template
# ============================================================================

def build_unified_template_bytes(
    *,
    expense_categories_tree: dict[str, Any],
    income_categories_tree: dict[str, Any],
    cards: list[str] | None = None,
) -> bytes:
    """
    Cria um Excel unificado com 4 abas:
    - Cartão: lançamentos de cartão de crédito
    - Débitos: gastos via débito/PIX/dinheiro  
    - Receitas: rendimentos
    - Contas Casa: contas fixas que entram no mês corrente
    """
    wb = Workbook()
    
    # Remove sheet padrão
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Criar as abas
    ws_cartao = wb.create_sheet("Cartão", 0)
    ws_debitos = wb.create_sheet("Débitos", 1)
    ws_receitas = wb.create_sheet("Receitas", 2)
    ws_contas = wb.create_sheet("Contas Casa", 3)
    ws_lists = wb.create_sheet("Listas", 4)
    
    # Processar categorias
    categories, sub_map, safe_map = _expense_categories(expense_categories_tree)
    income_cats = _income_categories(income_categories_tree)
    pessoas = ["Renan", "Aline", "Família"]
    cards_list = cards or ["Nubank", "C6", "XP", "Mercado Pago"]
    
    # ========================================================================
    # Aba Listas - dropdowns
    # ========================================================================
    
    # Coluna A: Categorias de despesa
    for i, cat in enumerate(categories, start=1):
        ws_lists.cell(row=i, column=1, value=cat)
    
    # Coluna B: Safe names para INDIRECT
    for i, cat in enumerate(categories, start=1):
        ws_lists.cell(row=i, column=2, value=safe_map[cat])
    
    # Colunas C+: Subcategorias por categoria
    base_col = 3
    for idx, cat in enumerate(categories):
        col = base_col + idx
        key = safe_map[cat]
        ws_lists.cell(row=1, column=col, value=key)
        subs = sub_map.get(cat, [])
        if not subs:
            subs = [""]
        for r_i, sub in enumerate(subs, start=2):
            ws_lists.cell(row=r_i, column=col, value=sub)
        
        # Named range para subcategorias
        col_letter = get_column_letter(col)
        dn = DefinedName(key, attr_text=f"Listas!${col_letter}$2:${col_letter}${1 + len(subs)}")
        _add_defined_name(wb, dn)
    
    # Named range para categorias
    if categories:
        dn = DefinedName("Categorias", attr_text=f"Listas!$A$1:$A${len(categories)}")
        _add_defined_name(wb, dn)
    
    # Coluna para Pessoas (longe das subcategorias)
    pessoa_col = 50
    for i, p in enumerate(pessoas, start=1):
        ws_lists.cell(row=i, column=pessoa_col, value=p)
    dn = DefinedName("Pessoas", attr_text=f"Listas!${get_column_letter(pessoa_col)}$1:${get_column_letter(pessoa_col)}${len(pessoas)}")
    _add_defined_name(wb, dn)
    
    # Coluna para Cartões
    cartao_col = 51
    for i, c in enumerate(cards_list, start=1):
        ws_lists.cell(row=i, column=cartao_col, value=c)
    dn = DefinedName("Cartoes", attr_text=f"Listas!${get_column_letter(cartao_col)}$1:${get_column_letter(cartao_col)}${len(cards_list)}")
    _add_defined_name(wb, dn)
    
    # Coluna para Categorias de Receita
    income_col = 52
    for i, c in enumerate(income_cats, start=1):
        ws_lists.cell(row=i, column=income_col, value=c)
    if income_cats:
        dn = DefinedName("CategoriasReceita", attr_text=f"Listas!${get_column_letter(income_col)}$1:${get_column_letter(income_col)}${len(income_cats)}")
        _add_defined_name(wb, dn)
    
    # Coluna para Contas da Casa
    contas_col = 53
    for i, c in enumerate(CONTAS_CASA, start=1):
        ws_lists.cell(row=i, column=contas_col, value=c)
    dn = DefinedName("ContasCasa", attr_text=f"Listas!${get_column_letter(contas_col)}$1:${get_column_letter(contas_col)}${len(CONTAS_CASA)}")
    _add_defined_name(wb, dn)
    
    ws_lists.sheet_state = "hidden"
    
    # ========================================================================
    # Aba Cartão
    # Nota: "Pago por Aline" indica que Aline pagou algo compartilhado no 
    # cartão dela. Isso gera um crédito a ela de valor/2 no acerto.
    # ========================================================================
    cartao_headers = [
        "Hash (oculto)",        # A - hidden
        "Data da compra",       # B
        "Data do vencimento",   # C
        "Categoria",            # D
        "Subcategoria",         # E
        "Cartão de crédito",    # F
        "Descrição",            # G
        "Valor (R$)",           # H
        "Pago por Aline",       # I - "X" se Aline pagou algo compartilhado (gera crédito a ela)
        "Reembolsável",         # J
        "Status",               # K
        "Notas",                # L
        "__chave (oculto)",     # M - hidden, for INDIRECT
        "Origin ID (oculto)",   # N - hidden, stable UUID
    ]
    ws_cartao.append(cartao_headers)
    ws_cartao.freeze_panes = "A2"
    ws_cartao.column_dimensions["A"].hidden = True
    ws_cartao.column_dimensions["M"].hidden = True
    ws_cartao.column_dimensions["N"].hidden = True
    
    # Data validations para Cartão
    dv_cat = DataValidation(type="list", formula1="=Categorias", allow_blank=True)
    dv_sub = DataValidation(type="list", formula1="=INDIRECT($M2)", allow_blank=True)
    dv_card = DataValidation(type="list", formula1="=Cartoes", allow_blank=True)
    dv_pago_aline = DataValidation(type="list", formula1='"X,"', allow_blank=True)  # X ou vazio
    dv_reemb = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Pago,Em aberto"', allow_blank=True)
    
    ws_cartao.add_data_validation(dv_cat)
    ws_cartao.add_data_validation(dv_sub)
    ws_cartao.add_data_validation(dv_card)
    ws_cartao.add_data_validation(dv_pago_aline)
    ws_cartao.add_data_validation(dv_reemb)
    ws_cartao.add_data_validation(dv_status)
    
    for r in range(2, 5001):
        dv_cat.add(f"D{r}")
        dv_sub.add(f"E{r}")
        dv_card.add(f"F{r}")
        dv_pago_aline.add(f"I{r}")
        dv_reemb.add(f"J{r}")
        dv_status.add(f"K{r}")
        # Formula para INDIRECT
        lookup_end = max(1, len(categories))
        ws_cartao.cell(row=r, column=13, value=f'=IFERROR(VLOOKUP($D{r},Listas!$A$1:$B${lookup_end},2,FALSE),"")')
    
    # Larguras
    for idx, w in enumerate([12, 14, 14, 25, 25, 18, 45, 14, 12, 12, 12, 30, 10, 10], start=1):
        ws_cartao.column_dimensions[get_column_letter(idx)].width = w
    
    # ========================================================================
    # Aba Débitos
    # ========================================================================
    debito_headers = [
        "Data",                 # A
        "Categoria",            # B
        "Subcategoria",         # C
        "Descrição",            # D
        "Valor (R$)",           # E
        "Pago por Aline",       # F - X se Aline pagou despesa compartilhada
        "Reembolsável",         # G
        "Notas",                # H
        "__chave (oculto)",     # I - hidden
        "Hash (oculto)",        # J - hidden
        "Origin ID (oculto)",   # K - hidden, stable UUID
    ]
    ws_debitos.append(debito_headers)
    ws_debitos.freeze_panes = "A2"
    ws_debitos.column_dimensions["I"].hidden = True
    ws_debitos.column_dimensions["J"].hidden = True
    ws_debitos.column_dimensions["K"].hidden = True
    
    dv_cat_d = DataValidation(type="list", formula1="=Categorias", allow_blank=True)
    dv_sub_d = DataValidation(type="list", formula1="=INDIRECT($I2)", allow_blank=True)
    dv_pago_aline_d = DataValidation(type="list", formula1='"X,"', allow_blank=True)
    dv_reemb_d = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    
    ws_debitos.add_data_validation(dv_cat_d)
    ws_debitos.add_data_validation(dv_sub_d)
    ws_debitos.add_data_validation(dv_pago_aline_d)
    ws_debitos.add_data_validation(dv_reemb_d)
    
    for r in range(2, 5001):
        dv_cat_d.add(f"B{r}")
        dv_sub_d.add(f"C{r}")
        dv_pago_aline_d.add(f"F{r}")
        dv_reemb_d.add(f"G{r}")
        lookup_end = max(1, len(categories))
        ws_debitos.cell(row=r, column=9, value=f'=IFERROR(VLOOKUP($B{r},Listas!$A$1:$B${lookup_end},2,FALSE),"")')
    
    for idx, w in enumerate([14, 25, 25, 45, 14, 12, 12, 30, 10, 10, 10], start=1):
        ws_debitos.column_dimensions[get_column_letter(idx)].width = w
    
    # ========================================================================
    # Aba Receitas
    # ========================================================================
    receita_headers = [
        "Data",                 # A
        "Categoria",            # B
        "Descrição",            # C
        "Valor (R$)",           # D
        "Notas",                # E
        "Hash (oculto)",        # F - hidden
        "Origin ID (oculto)",   # G - hidden, stable UUID
    ]
    ws_receitas.append(receita_headers)
    ws_receitas.freeze_panes = "A2"
    ws_receitas.column_dimensions["F"].hidden = True
    ws_receitas.column_dimensions["G"].hidden = True
    
    dv_cat_r = DataValidation(type="list", formula1="=CategoriasReceita", allow_blank=True) if income_cats else None
    
    if dv_cat_r:
        ws_receitas.add_data_validation(dv_cat_r)
    
    for r in range(2, 5001):
        if dv_cat_r:
            dv_cat_r.add(f"B{r}")
    
    for idx, w in enumerate([14, 25, 45, 14, 30, 10, 10], start=1):
        ws_receitas.column_dimensions[get_column_letter(idx)].width = w
    
    # ========================================================================
    # Aba Contas Casa
    # Nota: Contas da casa entram no mês em que são PAGAS (diferente de Débitos),
    # mas guardam "Mês Referência" para você saber de qual mês era a conta.
    # ========================================================================
    contas_headers = [
        "Mês Referência",       # A - ex: "2026-01" (mês ao qual a conta se refere)
        "Categoria",            # B - Categoria de despesa
        "Subcategoria",         # C - Subcategoria
        "Descrição",            # D
        "Valor (R$)",           # E
        "Pago por Aline",       # F - X se Aline pagou (senão, assume Renan)
        "Data Pagamento",       # G - data que define em qual acerto entra
        "Notas",                # H
        "__chave_categoria (oculto)",  # I - Para fórmula VLOOKUP
        "Hash (oculto)",        # J - hidden
        "Origin ID (oculto)",   # K - hidden, stable UUID
    ]
    ws_contas.append(contas_headers)
    ws_contas.freeze_panes = "A2"
    
    dv_cat = DataValidation(type="list", formula1="=Categorias", allow_blank=True)
    dv_sub = DataValidation(type="list", formula1="=INDIRECT($I2)", allow_blank=True)
    dv_pago_aline = DataValidation(type="list", formula1='"X,"', allow_blank=True)
    
    ws_contas.add_data_validation(dv_cat)
    ws_contas.add_data_validation(dv_sub)
    ws_contas.add_data_validation(dv_pago_aline)
    
    lookup_end = max(1, len(categories))
    for r in range(2, 5001):
        dv_cat.add(f"B{r}")
        dv_sub.add(f"C{r}")
        dv_pago_aline.add(f"F{r}")
        # Fórmula para __chave_categoria
        ws_contas.cell(row=r, column=9, value=f'=IFERROR(VLOOKUP($B{r},Listas!$A$1:$B${lookup_end},2,FALSE),"")')
    
    ws_contas.column_dimensions["I"].hidden = True
    ws_contas.column_dimensions["J"].hidden = True
    ws_contas.column_dimensions["K"].hidden = True

    for idx, w in enumerate([14, 28, 28, 35, 14, 14, 14, 20, 10, 10, 10], start=1):
        ws_contas.column_dimensions[get_column_letter(idx)].width = w
    
    # Salvar
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================================
# Ensure file exists
# ============================================================================

def ensure_unified_excel(
    path: Path,
    *,
    expense_categories_tree: dict[str, Any],
    income_categories_tree: dict[str, Any],
    cards: list[str] | None = None,
) -> Path:
    """Cria o Excel unificado se não existir."""
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(
            build_unified_template_bytes(
                expense_categories_tree=expense_categories_tree,
                income_categories_tree=income_categories_tree,
                cards=cards,
            )
        )
    return path


# ============================================================================
# Append rows to Cartão sheet
# ============================================================================

def append_credit_card_rows(
    path: Path,
    *,
    rows: Iterable[dict[str, Any]],
    expense_categories_tree: dict[str, Any],
    income_categories_tree: dict[str, Any],
    cards: list[str] | None = None,
) -> int:
    """
    Adiciona linhas de cartão de crédito na aba Cartão.
    Deduplicação por row_hash.
    """
    ensure_unified_excel(
        path,
        expense_categories_tree=expense_categories_tree,
        income_categories_tree=income_categories_tree,
        cards=cards,
    )
    
    wb = load_workbook(path)
    ws = wb["Cartão"] if "Cartão" in wb.sheetnames else wb.sheetnames[0]
    
    # Map headers
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[normalize_str(v)] = c
    
    hash_col = _find_col(headers, ("hash (oculto)", "hash", "row_hash")) or 1
    txn_col = _find_col(headers, ("data da compra", "txn_date")) or 2
    due_col = _find_col(headers, ("data do vencimento", "vencimento")) or 3
    cat_col = _find_col(headers, ("categoria",)) or 4
    sub_col = _find_col(headers, ("subcategoria",)) or 5
    card_col = _find_col(headers, ("cartao de credito", "cartão de crédito")) or 6
    desc_col = _find_col(headers, ("descricao", "descrição")) or 7
    amount_col = _find_col(headers, ("valor (r$)", "valor")) or 8
    pago_aline_col = _find_col(headers, ("pago por aline",)) or 9
    reemb_col = _find_col(headers, ("reembolsavel", "reembolsável")) or 10
    status_col = _find_col(headers, ("status",)) or 11
    notes_col = _find_col(headers, ("notas",)) or 12
    origin_id_col = _find_col(headers, ("origin id (oculto)", "origin_id"))

    # Existing hashes and origin_ids — used for deduplication
    existing_hashes: set[str] = set()
    existing_origin_ids: set[str] = set()
    last_row = 1
    for r in range(2, ws.max_row + 1):
        rh = str(ws.cell(row=r, column=hash_col).value or "").strip()
        if rh:
            existing_hashes.add(rh)
            last_row = r
            if origin_id_col:
                oid = str(ws.cell(row=r, column=origin_id_col).value or "").strip()
                if oid:
                    existing_origin_ids.add(oid)
            continue
        # Avoid jumping to very large row numbers because of stray values in a single cell.
        # Only treat row as "occupied" when it looks like a real credit-card entry.
        txn_v = ws.cell(row=r, column=txn_col).value
        due_v = ws.cell(row=r, column=due_col).value
        desc_v = ws.cell(row=r, column=desc_col).value
        card_v = ws.cell(row=r, column=card_col).value
        amount_v = ws.cell(row=r, column=amount_col).value
        row_has_core_data = (
            (txn_v not in (None, "") or due_v not in (None, ""))
            and desc_v not in (None, "")
            and card_v not in (None, "")
            and amount_v not in (None, "")
        )
        if row_has_core_data:
            last_row = r

    appended = 0
    today = date.today()
    next_row = last_row + 1

    for row in rows:
        rh = str(row.get("row_hash") or "").strip()
        oid = str(row.get("origin_id") or "").strip()

        # Skip if already in sheet — check by origin_id first, then row_hash
        if oid and oid in existing_origin_ids:
            continue
        if rh and rh in existing_hashes:
            continue
        if not rh:
            continue

        txn_date = parse_date(row.get("txn_date"))
        due_date = parse_date(row.get("statement_due_date") or row.get("cash_date"))
        if txn_date is None or due_date is None:
            continue

        ws.cell(row=next_row, column=hash_col, value=rh)
        ws.cell(row=next_row, column=txn_col, value=txn_date)
        ws.cell(row=next_row, column=due_col, value=due_date)
        ws.cell(row=next_row, column=cat_col, value=row.get("category") or "")
        ws.cell(row=next_row, column=sub_col, value=row.get("subcategory") or "")
        ws.cell(row=next_row, column=card_col, value=row.get("account") or "")
        ws.cell(row=next_row, column=desc_col, value=row.get("description") or "")

        amt = row.get("amount")
        if amt is not None:
            ws.cell(row=next_row, column=amount_col, value=float(amt))

        # Pago por Aline fica vazio por padrão - usuário preenche manualmente
        ws.cell(row=next_row, column=pago_aline_col, value="")
        ws.cell(row=next_row, column=reemb_col, value="Sim" if row.get("reimbursable") else "Não")
        ws.cell(row=next_row, column=status_col, value="Pago" if due_date <= today else "Em aberto")
        ws.cell(row=next_row, column=notes_col, value=row.get("notes") or "")
        if origin_id_col and oid:
            ws.cell(row=next_row, column=origin_id_col, value=oid)

        existing_hashes.add(rh)
        if oid:
            existing_origin_ids.add(oid)
        appended += 1
        next_row += 1
    
    if appended:
        wb.save(path)
    return appended


def _last_used_row(ws, *, key_cols: tuple[int, ...]) -> int:
    last = 1
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in key_cols):
            last = r
    return last


def append_transactions_to_unified(
    path: Path,
    *,
    rows: Iterable[dict[str, Any]],
    expense_categories_tree: dict[str, Any],
    income_categories_tree: dict[str, Any],
    cards: list[str] | None = None,
) -> dict[str, int]:
    """
    Adiciona lançamentos no Excel unificado, em suas respectivas abas.
    Espera rows no formato de transações do banco (txn_date/cash_date/amount/payment_method...).
    """
    all_rows = list(rows)
    if not all_rows:
        return {"credit_card": 0, "debit": 0, "income": 0, "household": 0}

    ensure_unified_excel(
        path,
        expense_categories_tree=expense_categories_tree,
        income_categories_tree=income_categories_tree,
        cards=cards,
    )

    out = {"credit_card": 0, "debit": 0, "income": 0, "household": 0}

    cc_rows = [r for r in all_rows if str(r.get("payment_method") or "") == "credit_card"]
    if cc_rows:
        out["credit_card"] = append_credit_card_rows(
            path,
            rows=cc_rows,
            expense_categories_tree=expense_categories_tree,
            income_categories_tree=income_categories_tree,
            cards=cards,
        )

    non_cc_rows = [r for r in all_rows if str(r.get("payment_method") or "") != "credit_card"]
    if not non_cc_rows:
        return out

    wb = load_workbook(path)
    ws_debitos = wb["Débitos"] if "Débitos" in wb.sheetnames else None
    ws_receitas = wb["Receitas"] if "Receitas" in wb.sheetnames else None
    ws_contas = wb["Contas Casa"] if "Contas Casa" in wb.sheetnames else None

    next_deb = (_last_used_row(ws_debitos, key_cols=(1, 4, 5)) + 1) if ws_debitos else 2
    next_rec = (_last_used_row(ws_receitas, key_cols=(1, 3, 4)) + 1) if ws_receitas else 2
    next_house = (_last_used_row(ws_contas, key_cols=(1, 2, 4, 5, 7)) + 1) if ws_contas else 2

    for row in non_cc_rows:
        pm = str(row.get("payment_method") or "").strip()
        txn_dt = parse_date(row.get("txn_date"))
        cash_dt = parse_date(row.get("cash_date"))
        amount = row.get("amount")
        if amount is None:
            continue
        amount_f = float(amount)
        description = str(row.get("description") or "").strip()
        category = str(row.get("category") or "").strip()
        subcategory = str(row.get("subcategory") or "").strip()
        person = str(row.get("person") or "").strip()
        notes = str(row.get("notes") or "").strip()

        if pm == "debit" and ws_debitos is not None:
            if txn_dt is None or not description:
                continue
            ws_debitos.cell(row=next_deb, column=1, value=txn_dt)
            ws_debitos.cell(row=next_deb, column=2, value=category)
            ws_debitos.cell(row=next_deb, column=3, value=subcategory)
            ws_debitos.cell(row=next_deb, column=4, value=description)
            ws_debitos.cell(row=next_deb, column=5, value=abs(amount_f))
            ws_debitos.cell(row=next_deb, column=6, value="X" if normalize_str(person) == "aline" else "")
            ws_debitos.cell(row=next_deb, column=7, value="Sim" if bool(row.get("reimbursable")) else "Não")
            ws_debitos.cell(row=next_deb, column=8, value=notes)
            out["debit"] += 1
            next_deb += 1

        elif pm == "income" and ws_receitas is not None:
            if txn_dt is None or not description:
                continue
            ws_receitas.cell(row=next_rec, column=1, value=txn_dt)
            ws_receitas.cell(row=next_rec, column=2, value=category)
            ws_receitas.cell(row=next_rec, column=3, value=description)
            ws_receitas.cell(row=next_rec, column=4, value=abs(amount_f))
            ws_receitas.cell(row=next_rec, column=5, value=notes)
            out["income"] += 1
            next_rec += 1

        elif pm == "household" and ws_contas is not None:
            if not description:
                continue
            payment_dt = cash_dt or txn_dt
            if payment_dt is None:
                continue
            ref_month = str(row.get("reference") or "").strip() or payment_dt.strftime("%Y-%m")
            ws_contas.cell(row=next_house, column=1, value=ref_month)
            ws_contas.cell(row=next_house, column=2, value=category)
            ws_contas.cell(row=next_house, column=3, value=subcategory)
            ws_contas.cell(row=next_house, column=4, value=description)
            ws_contas.cell(row=next_house, column=5, value=abs(amount_f))
            ws_contas.cell(row=next_house, column=6, value="X" if normalize_str(person) == "aline" else "")
            ws_contas.cell(row=next_house, column=7, value=payment_dt)
            ws_contas.cell(row=next_house, column=8, value=notes)
            out["household"] += 1
            next_house += 1

    if any(v > 0 for v in out.values()):
        wb.save(path)
    return out


# ============================================================================
# Read from Excel sheets
# ============================================================================

def read_cartao_sheet(path: Path) -> list[dict[str, Any]]:
    """Lê a aba Cartão e retorna lista de dicts."""
    if not path.exists():
        return []
    
    wb = load_workbook(path, data_only=True)
    if "Cartão" not in wb.sheetnames:
        return []
    
    ws = wb["Cartão"]
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[normalize_str(v)] = c
    
    hash_col = _find_col(headers, ("hash (oculto)", "hash")) or 1
    txn_col = _find_col(headers, ("data da compra",)) or 2
    due_col = _find_col(headers, ("data do vencimento",)) or 3
    cat_col = _find_col(headers, ("categoria",)) or 4
    sub_col = _find_col(headers, ("subcategoria",)) or 5
    card_col = _find_col(headers, ("cartao de credito", "cartão de crédito")) or 6
    desc_col = _find_col(headers, ("descricao", "descrição")) or 7
    amount_col = _find_col(headers, ("valor (r$)", "valor")) or 8
    pago_aline_col = _find_col(headers, ("pago por aline", "pessoa")) or 9
    reemb_col = _find_col(headers, ("reembolsavel", "reembolsável")) or 10
    status_col = _find_col(headers, ("status",)) or 11
    notes_col = _find_col(headers, ("notas",)) or 12
    origin_id_col = _find_col(headers, ("origin id (oculto)", "origin_id"))

    rows = []
    for r in range(2, ws.max_row + 1):
        rh = ws.cell(row=r, column=hash_col).value
        desc = ws.cell(row=r, column=desc_col).value
        if not rh and not desc:
            continue

        amt = ws.cell(row=r, column=amount_col).value
        try:
            amount = float(amt) if amt else None
        except Exception:
            amount = parse_brl_number(amt)

        reemb_val = str(ws.cell(row=r, column=reemb_col).value or "").strip().lower()
        pago_aline_val = str(ws.cell(row=r, column=pago_aline_col).value or "").strip().upper()
        origin_id = str(ws.cell(row=r, column=origin_id_col).value or "").strip() if origin_id_col else ""

        rows.append({
            "row_hash": str(rh or "").strip(),
            "origin_id": origin_id,
            "txn_date": parse_date(ws.cell(row=r, column=txn_col).value),
            "due_date": parse_date(ws.cell(row=r, column=due_col).value),
            "category": str(ws.cell(row=r, column=cat_col).value or "").strip(),
            "subcategory": str(ws.cell(row=r, column=sub_col).value or "").strip(),
            "account": str(ws.cell(row=r, column=card_col).value or "").strip(),
            "description": str(desc or "").strip(),
            "amount": amount,
            "pago_por_aline": pago_aline_val == "X",
            "reimbursable": reemb_val in ("sim", "s", "yes", "1", "true"),
            "status": str(ws.cell(row=r, column=status_col).value or "").strip(),
            "notes": str(ws.cell(row=r, column=notes_col).value or "").strip(),
        })
    
    return rows


def update_credit_card_status(
    path: Path,
    *,
    account: str,
    due_date: date,
    status: str,
) -> int:
    """
    Atualiza a coluna "Status" da aba Cartão para um cartão + vencimento.
    Retorna o número de linhas atualizadas.
    """
    if not path.exists():
        return 0
    wb = load_workbook(path)
    if "Cartão" not in wb.sheetnames:
        return 0

    ws = wb["Cartão"]
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[normalize_str(v)] = c

    due_col = _find_col(headers, ("data do vencimento", "vencimento")) or 3
    card_col = _find_col(headers, ("cartao de credito", "cartão de crédito")) or 6
    status_col = _find_col(headers, ("status",)) or 11

    target_account = normalize_str(account)
    updated = 0
    for r in range(2, ws.max_row + 1):
        acc_val = normalize_str(str(ws.cell(row=r, column=card_col).value or ""))
        if not acc_val or acc_val != target_account:
            continue
        due_val = parse_date(ws.cell(row=r, column=due_col).value)
        if due_val is None or due_val != due_date:
            continue
        ws.cell(row=r, column=status_col, value=str(status).strip())
        updated += 1

    if updated:
        wb.save(path)
    return updated


def update_credit_card_categories(
    path: Path,
    *,
    updates: Iterable[dict[str, Any]],
) -> tuple[int, int]:
    """
    Atualiza Categoria/Subcategoria/Reembolsável na aba "Cartão" por row_hash.
    Retorna (updated_count, missing_count).
    """
    if not path.exists():
        return 0, 0
    wb = load_workbook(path)
    if "Cartão" not in wb.sheetnames:
        return 0, 0

    ws = wb["Cartão"]
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[normalize_str(v)] = c

    hash_col = _find_col(headers, ("hash (oculto)", "hash", "row_hash")) or 1
    cat_col = _find_col(headers, ("categoria",)) or 4
    sub_col = _find_col(headers, ("subcategoria",)) or 5
    reemb_col = _find_col(headers, ("reembolsavel", "reembolsável", "reimbursable")) or 10

    row_by_hash: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        rh = str(ws.cell(row=r, column=hash_col).value or "").strip()
        if rh and rh not in row_by_hash:
            row_by_hash[rh] = r

    updated = 0
    missing = 0
    touched = False

    for u in updates:
        rh = str(u.get("row_hash") or "").strip()
        if not rh:
            continue
        row = row_by_hash.get(rh)
        if row is None:
            missing += 1
            continue

        changed = False

        if "category" in u:
            new_cat = (str(u.get("category") or "").strip() or "")
            old_cat = str(ws.cell(row=row, column=cat_col).value or "").strip()
            if new_cat != old_cat:
                ws.cell(row=row, column=cat_col, value=new_cat)
                changed = True

        if "subcategory" in u:
            new_sub = (str(u.get("subcategory") or "").strip() or "")
            old_sub = str(ws.cell(row=row, column=sub_col).value or "").strip()
            if new_sub != old_sub:
                ws.cell(row=row, column=sub_col, value=new_sub)
                changed = True

        if "reimbursable" in u:
            new_reemb = "Sim" if bool(u.get("reimbursable")) else "Não"
            old_reemb = str(ws.cell(row=row, column=reemb_col).value or "").strip()
            if new_reemb != old_reemb:
                ws.cell(row=row, column=reemb_col, value=new_reemb)
                changed = True

        if changed:
            updated += 1
            touched = True

    if touched:
        wb.save(path)
    return updated, missing


def read_debitos_sheet(path: Path) -> list[dict[str, Any]]:
    """Lê a aba Débitos e retorna lista de dicts."""
    if not path.exists():
        return []
    
    wb = load_workbook(path, data_only=True)
    if "Débitos" not in wb.sheetnames:
        return []
    
    ws = wb["Débitos"]
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[normalize_str(v)] = c
    
    hash_col = _find_col(headers, ("hash (oculto)", "hash", "row_hash"))
    origin_id_col = _find_col(headers, ("origin id (oculto)", "origin_id"))
    date_col = _find_col(headers, ("data",)) or 1
    cat_col = _find_col(headers, ("categoria",)) or 2
    sub_col = _find_col(headers, ("subcategoria",)) or 3
    desc_col = _find_col(headers, ("descricao", "descrição")) or 4
    amount_col = _find_col(headers, ("valor (r$)", "valor")) or 5
    pago_aline_col = _find_col(headers, ("pago por aline",)) or 6
    reemb_col = _find_col(headers, ("reembolsavel", "reembolsável")) or 7
    notes_col = _find_col(headers, ("notas",)) or 8

    rows = []
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(row=r, column=desc_col).value
        dt = ws.cell(row=r, column=date_col).value
        if not desc and not dt:
            continue

        amt = ws.cell(row=r, column=amount_col).value
        try:
            amount = float(amt) if amt else None
        except Exception:
            amount = parse_brl_number(amt)

        reemb_val = str(ws.cell(row=r, column=reemb_col).value or "").strip().lower()
        pago_aline_val = str(ws.cell(row=r, column=pago_aline_col).value or "").strip().upper()
        origin_id = str(ws.cell(row=r, column=origin_id_col).value or "").strip() if origin_id_col else ""

        rows.append({
            "row_hash": str(ws.cell(row=r, column=hash_col).value or "").strip() if hash_col else "",
            "origin_id": origin_id,
            "date": parse_date(dt),
            "category": str(ws.cell(row=r, column=cat_col).value or "").strip(),
            "subcategory": str(ws.cell(row=r, column=sub_col).value or "").strip(),
            "description": str(desc or "").strip(),
            "amount": amount,
            "pago_por_aline": pago_aline_val == "X",
            "reimbursable": reemb_val in ("sim", "s", "yes", "1", "true"),
            "notes": str(ws.cell(row=r, column=notes_col).value or "").strip(),
        })

    return rows


def read_receitas_sheet(path: Path) -> list[dict[str, Any]]:
    """Lê a aba Receitas e retorna lista de dicts."""
    if not path.exists():
        return []
    
    wb = load_workbook(path, data_only=True)
    if "Receitas" not in wb.sheetnames:
        return []
    
    ws = wb["Receitas"]
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[normalize_str(v)] = c
    
    hash_col = _find_col(headers, ("hash (oculto)", "hash", "row_hash"))
    origin_id_col = _find_col(headers, ("origin id (oculto)", "origin_id"))
    date_col = _find_col(headers, ("data",)) or 1
    cat_col = _find_col(headers, ("categoria",)) or 2
    desc_col = _find_col(headers, ("descricao", "descrição")) or 3
    amount_col = _find_col(headers, ("valor (r$)", "valor")) or 4
    recebido_aline_col = _find_col(headers, ("recebido por aline",))
    person_col = _find_col(headers, ("pessoa", "person"))
    notes_col = _find_col(headers, ("notas",)) or 5

    rows = []
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(row=r, column=desc_col).value
        dt = ws.cell(row=r, column=date_col).value
        if not desc and not dt:
            continue

        amt = ws.cell(row=r, column=amount_col).value
        try:
            amount = float(amt) if amt else None
        except Exception:
            amount = parse_brl_number(amt)

        recebido_aline_val = (
            str(ws.cell(row=r, column=recebido_aline_col).value or "").strip().upper()
            if recebido_aline_col
            else ""
        )
        person_val = str(ws.cell(row=r, column=person_col).value or "").strip() if person_col else ""
        person = person_val or ("Aline" if recebido_aline_val == "X" else "")
        origin_id = str(ws.cell(row=r, column=origin_id_col).value or "").strip() if origin_id_col else ""

        rows.append({
            "row_hash": str(ws.cell(row=r, column=hash_col).value or "").strip() if hash_col else "",
            "origin_id": origin_id,
            "date": parse_date(dt),
            "category": str(ws.cell(row=r, column=cat_col).value or "").strip(),
            "description": str(desc or "").strip(),
            "amount": amount,
            "person": person or None,
            "recebido_por_aline": recebido_aline_val == "X",
            "notes": str(ws.cell(row=r, column=notes_col).value or "").strip(),
        })

    return rows


def read_contas_casa_sheet(path: Path) -> list[dict[str, Any]]:
    """
    Lê a aba Contas Casa e retorna lista de dicts.
    
    Nota: Contas da casa entram no mês em que foram pagas (via Data Pagamento),
    e o "Mês Referência" serve só para identificar de qual mês era a conta.
    """
    if not path.exists():
        return []
    
    wb = load_workbook(path, data_only=True)
    if "Contas Casa" not in wb.sheetnames:
        return []
    
    ws = wb["Contas Casa"]
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[normalize_str(v)] = c
    
    hash_col = _find_col(headers, ("hash (oculto)", "hash", "row_hash"))
    origin_id_col = _find_col(headers, ("origin id (oculto)", "origin_id"))
    ref_col = _find_col(headers, ("mes referencia", "mês referência")) or 1
    cat_col = _find_col(headers, ("categoria", "category")) or 2
    subcat_col = _find_col(headers, ("subcategoria", "subcategory")) or 3
    desc_col = _find_col(headers, ("descricao", "descrição")) or 4
    amount_col = _find_col(headers, ("valor (r$)", "valor")) or 5
    pago_aline_col = _find_col(headers, ("pago por aline",))
    quem_col = _find_col(headers, ("pago por", "quem pagou", "pessoa", "person"))
    pag_col = _find_col(headers, ("data pagamento",)) or 7
    notes_col = _find_col(headers, ("notas",)) or 8

    rows = []
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(row=r, column=ref_col).value
        category = ws.cell(row=r, column=cat_col).value
        if not ref and not category:
            continue

        amt = ws.cell(row=r, column=amount_col).value
        try:
            amount = float(amt) if amt else None
        except Exception:
            amount = parse_brl_number(amt)

        ref_month = _parse_reference_month(ref)

        pago_aline_val = (
            str(ws.cell(row=r, column=pago_aline_col).value or "").strip().upper()
            if pago_aline_col
            else ""
        )
        paid_by = None
        if pago_aline_col:
            paid_by = "Aline" if pago_aline_val == "X" else "Renan"
        elif quem_col:
            paid_by = str(ws.cell(row=r, column=quem_col).value or "").strip() or None

        origin_id = str(ws.cell(row=r, column=origin_id_col).value or "").strip() if origin_id_col else ""

        rows.append({
            "row_hash": str(ws.cell(row=r, column=hash_col).value or "").strip() if hash_col else "",
            "origin_id": origin_id,
            "reference_month": ref_month,
            "category": str(category or "").strip(),
            "subcategory": str(ws.cell(row=r, column=subcat_col).value or "").strip() or None,
            "description": str(ws.cell(row=r, column=desc_col).value or "").strip(),
            "amount": amount,
            "paid_by": paid_by,
            "pago_por_aline": pago_aline_val == "X" if pago_aline_col else None,
            "payment_date": parse_date(ws.cell(row=r, column=pag_col).value),
            "notes": str(ws.cell(row=r, column=notes_col).value or "").strip(),
        })
    
    return rows

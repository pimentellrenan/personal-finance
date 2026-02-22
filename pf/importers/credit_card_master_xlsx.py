from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pf.db import now_iso
from pf.utils import normalize_str, parse_brl_number, parse_date, sha256_text


@dataclass(frozen=True)
class CreditCardMasterXlsx:
    rows: list[dict[str, Any]]
    hash_updates: dict[int, str]  # excel row index (1-based) -> hash
    hash_col: int  # excel column index (1-based)
    sheet_name: str


def _find_col(headers: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    # Exact match
    for cand in candidates:
        key = normalize_str(cand)
        if key in headers:
            return headers[key]
    # Partial match
    for cand in candidates:
        key = normalize_str(cand)
        for h, idx in headers.items():
            if key and key in h:
                return idx
    return None


def read_credit_card_master_xlsx(
    path: Path,
    *,
    source_hash: str,
    source_file: str | None = None,
    card_owner_by_name: dict[str, str] | None = None,
) -> CreditCardMasterXlsx:
    wb = load_workbook(path, data_only=True)
    sheet_name = "Cartão" if "Cartão" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        key = normalize_str(v)
        if key:
            headers[key] = c

    hash_col = _find_col(headers, ("hash (oculto)", "hash", "row_hash", "id"))
    origin_id_col = _find_col(headers, ("origin id (oculto)", "origin_id", "origin id"))

    txn_col = _find_col(headers, ("data da compra", "data compra", "txn_date", "data"))
    due_col = _find_col(headers, ("data do vencimento", "vencimento", "statement_due_date"))
    cat_col = _find_col(headers, ("categoria", "category"))
    sub_col = _find_col(headers, ("subcategoria", "sub-category", "sub_category", "subcategory"))
    card_col = _find_col(headers, ("cartao de credito", "cartão de crédito", "cartao", "cartão", "account"))
    desc_col = _find_col(headers, ("descricao", "descrição", "description"))
    amount_col = _find_col(headers, ("valor", "valor (r$)", "valor (em r$)", "amount"))
    status_col = _find_col(headers, ("status (cartao pago ou nao)", "status"))
    reemb_col = _find_col(headers, ("reembolsavel", "reembolsável", "reimbursable"))
    notes_col = _find_col(headers, ("notas", "notes"))
    # NOTE: In the unified workbook, there's a boolean flag column "Pago por Aline" (X/blank),
    # which is not the same as a free-text "Pessoa" column. Avoid matching generic "pago por".
    person_col = _find_col(headers, ("pessoa", "person", "quem pagou", "portador", "titular"))
    pago_por_aline_col = _find_col(headers, ("pago por aline",))

    if txn_col is None or due_col is None or desc_col is None or amount_col is None or card_col is None:
        raise ValueError(
            "XLSX de cartão não reconhecido. Colunas esperadas: "
            "`Data da compra`, `Data do vencimento`, `Cartão de crédito`, `Descrição`, `Valor`."
        )

    created = now_iso()
    src_file = source_file or str(path)

    rows: list[dict[str, Any]] = []
    # Read-only import: never writes back hashes to the workbook.
    hash_updates: dict[int, str] = {}

    for r in range(2, ws.max_row + 1):
        rh = ""
        if hash_col:
            rh_raw = ws.cell(row=r, column=hash_col).value
            rh = str(rh_raw or "").strip()
        origin_id = None
        if origin_id_col:
            oid_raw = ws.cell(row=r, column=origin_id_col).value
            origin_id = str(oid_raw or "").strip() or None

        txn_dt = parse_date(ws.cell(row=r, column=txn_col).value)
        due_dt = parse_date(ws.cell(row=r, column=due_col).value)
        description = str(ws.cell(row=r, column=desc_col).value or "").strip()
        amount_raw = parse_brl_number(ws.cell(row=r, column=amount_col).value)
        card = str(ws.cell(row=r, column=card_col).value or "").strip()

        # Skip completely empty rows
        if txn_dt is None and due_dt is None and not description and amount_raw is None and not card and not rh:
            continue

        if txn_dt is None or due_dt is None or not description or amount_raw is None or not card:
            # Partial row filled. Ignore silently to avoid injecting bad rows; user can fix in Excel.
            continue

        category = str(ws.cell(row=r, column=cat_col).value or "").strip() if cat_col else ""
        subcategory = str(ws.cell(row=r, column=sub_col).value or "").strip() if sub_col else ""

        status = ""
        if status_col:
            status = str(ws.cell(row=r, column=status_col).value or "").strip()

        person = None
        if person_col:
            person_val = str(ws.cell(row=r, column=person_col).value or "").strip()
            person = person_val if person_val else None
        if person is None and pago_por_aline_col:
            pago_val = str(ws.cell(row=r, column=pago_por_aline_col).value or "").strip().upper()
            if pago_val == "X":
                person = "Aline"
        if person is None and card_owner_by_name:
            person = card_owner_by_name.get(card) or None

        # Preserve sign: expenses < 0, credits/refunds > 0.
        amount = float(amount_raw)

        reimbursable = 0
        if reemb_col:
            reemb_val = str(ws.cell(row=r, column=reemb_col).value or "").strip().lower()
            reimbursable = 1 if reemb_val in ("sim", "s", "yes", "1", "true") else 0

        notes = None
        notes_cell = str(ws.cell(row=r, column=notes_col).value or "").strip() if notes_col else ""
        status_note = f"Status: {status}" if status else ""
        if notes_cell and status_note:
            notes = f"{notes_cell} | {status_note}"
        elif notes_cell:
            notes = notes_cell
        elif status_note:
            notes = status_note

        if not rh:
            rh = sha256_text(
                "|".join(
                    [
                        "excel_credit_card",
                        txn_dt.isoformat(),
                        due_dt.isoformat(),
                        f"{amount:.2f}",
                        normalize_str(description),
                        normalize_str(card),
                    ]
                )
            )

        rows.append(
            {
                "origin_id": origin_id,
                "row_hash": rh,
                "txn_date": txn_dt.isoformat(),
                "cash_date": due_dt.isoformat(),
                "amount": amount,
                "description": description,
                "group_name": None,
                "category": category or None,
                "subcategory": subcategory or None,
                "payment_method": "credit_card",
                "account": card,
                "source": "excel_credit_card",
                "statement_closing_date": None,
                "statement_due_date": due_dt.isoformat(),
                "person": person,
                "reimbursable": reimbursable,
                "reference": None,
                "notes": notes,
                "source_file": src_file,
                "source_hash": source_hash,
                "external_id": None,
                "created_at": created,
                "updated_at": created,
            }
        )

    return CreditCardMasterXlsx(
        rows=rows,
        hash_updates=hash_updates,
        hash_col=int(hash_col or 0),
        sheet_name=sheet_name,
    )


def write_credit_card_hashes(path: Path, *, sheet_name: str, hash_col: int, updates: dict[int, str]) -> None:
    if not updates:
        return
    wb = load_workbook(path)
    ws = wb[sheet_name]
    for row_idx, rh in updates.items():
        ws.cell(row=int(row_idx), column=int(hash_col)).value = str(rh)
    wb.save(path)

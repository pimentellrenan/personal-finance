"""
Importer unificado para o arquivo lancamentos.xlsx com múltiplas abas:
- Cartão de Crédito
- Débitos  
- Receitas
- Contas da Casa
"""
from __future__ import annotations

from pathlib import Path
from typing import Any
from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook

from pf.db import now_iso
from pf.utils import normalize_str, parse_brl_number, parse_date, sha256_text


@dataclass
class UnifiedImportResult:
    """Resultado da importação do arquivo unificado"""
    credit_card: list[dict[str, Any]]
    debit: list[dict[str, Any]]
    income: list[dict[str, Any]]
    household: list[dict[str, Any]]  # Contas da Casa


def _find_col(headers: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    """Encontra coluna por nome normalizado"""
    for cand in candidates:
        key = normalize_str(cand)
        if key in headers:
            return headers[key]
    for cand in candidates:
        key = normalize_str(cand)
        for h, idx in headers.items():
            if key and key in h:
                return idx
    return None


def _get_headers(ws) -> dict[str, int]:
    """Extrai headers normalizados da primeira linha"""
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        key = normalize_str(v)
        if key:
            headers[key] = c
    return headers


def import_unified_xlsx(
    path: Path,
    *,
    source_hash: str,
    source_file: str | None = None,
) -> UnifiedImportResult:
    """
    Importa todas as abas do arquivo unificado.
    """
    wb = load_workbook(path, data_only=True)
    source_file = source_file or str(path)
    created = now_iso()
    
    result = UnifiedImportResult(
        credit_card=[],
        debit=[],
        income=[],
        household=[],
    )
    
    # Import each sheet if present
    if "Cartão de Crédito" in wb.sheetnames:
        result.credit_card = _import_credit_card_sheet(
            wb["Cartão de Crédito"], source_hash, source_file, created
        )
    
    if "Débitos" in wb.sheetnames:
        result.debit = _import_debit_sheet(
            wb["Débitos"], source_hash, source_file, created
        )
    
    if "Receitas" in wb.sheetnames:
        result.income = _import_income_sheet(
            wb["Receitas"], source_hash, source_file, created
        )
    
    if "Contas da Casa" in wb.sheetnames:
        result.household = _import_household_sheet(
            wb["Contas da Casa"], source_hash, source_file, created
        )
    
    return result


def _import_credit_card_sheet(
    ws, source_hash: str, source_file: str, created: str
) -> list[dict[str, Any]]:
    """Importa aba Cartão de Crédito"""
    headers = _get_headers(ws)
    
    txn_col = _find_col(headers, ("data da compra", "data compra", "data"))
    due_col = _find_col(headers, ("data do vencimento", "vencimento"))
    cat_col = _find_col(headers, ("categoria", "category"))
    sub_col = _find_col(headers, ("subcategoria", "sub-category"))
    card_col = _find_col(headers, ("cartao de credito", "cartao", "cartão"))
    desc_col = _find_col(headers, ("descricao", "descrição", "description"))
    amount_col = _find_col(headers, ("valor", "amount"))
    status_col = _find_col(headers, ("status",))
    person_col = _find_col(headers, ("pessoa", "person", "quem pagou"))
    
    if not all([txn_col, desc_col, amount_col]):
        return []
    
    rows: list[dict[str, Any]] = []
    
    for r in range(2, ws.max_row + 1):
        txn_dt = parse_date(ws.cell(row=r, column=txn_col).value) if txn_col else None
        if txn_dt is None:
            continue
        
        description = str(ws.cell(row=r, column=desc_col).value or "").strip()
        if not description:
            continue
        
        amount_raw = parse_brl_number(ws.cell(row=r, column=amount_col).value)
        if amount_raw is None:
            continue
        amount = -abs(float(amount_raw))
        
        due_dt = parse_date(ws.cell(row=r, column=due_col).value) if due_col else None
        category = str(ws.cell(row=r, column=cat_col).value or "").strip() or None if cat_col else None
        subcategory = str(ws.cell(row=r, column=sub_col).value or "").strip() or None if sub_col else None
        card = str(ws.cell(row=r, column=card_col).value or "").strip() or None if card_col else None
        person = str(ws.cell(row=r, column=person_col).value or "").strip() or None if person_col else None
        
        row_hash = sha256_text(
            "|".join([
                "credit_card",
                txn_dt.isoformat(),
                f"{amount:.2f}",
                description,
                card or "",
            ])
        )
        
        rows.append({
            "row_hash": row_hash,
            "txn_date": txn_dt.isoformat(),
            "cash_date": due_dt.isoformat() if due_dt else txn_dt.isoformat(),
            "amount": amount,
            "description": description,
            "group_name": None,
            "category": category,
            "subcategory": subcategory,
            "payment_method": "credit_card",
            "account": card,
            "source": "unified_xlsx",
            "statement_closing_date": None,
            "statement_due_date": due_dt.isoformat() if due_dt else None,
            "person": person,
            "reimbursable": 0,
            "reference": None,
            "notes": None,
            "source_file": source_file,
            "source_hash": source_hash,
            "external_id": None,
            "created_at": created,
            "updated_at": created,
        })
    
    return rows


def _import_debit_sheet(
    ws, source_hash: str, source_file: str, created: str
) -> list[dict[str, Any]]:
    """Importa aba Débitos"""
    headers = _get_headers(ws)
    
    date_col = _find_col(headers, ("data", "date"))
    cat_col = _find_col(headers, ("categoria", "category"))
    sub_col = _find_col(headers, ("subcategoria", "sub-category"))
    desc_col = _find_col(headers, ("descricao", "descrição", "description"))
    amount_col = _find_col(headers, ("valor", "amount"))
    person_col = _find_col(headers, ("pessoa", "person", "quem pagou"))
    
    if not all([date_col, desc_col, amount_col]):
        return []
    
    rows: list[dict[str, Any]] = []
    
    for r in range(2, ws.max_row + 1):
        txn_dt = parse_date(ws.cell(row=r, column=date_col).value) if date_col else None
        if txn_dt is None:
            continue
        
        description = str(ws.cell(row=r, column=desc_col).value or "").strip()
        if not description:
            continue
        
        amount_raw = parse_brl_number(ws.cell(row=r, column=amount_col).value)
        if amount_raw is None:
            continue
        amount = -abs(float(amount_raw))
        
        category = str(ws.cell(row=r, column=cat_col).value or "").strip() or None if cat_col else None
        subcategory = str(ws.cell(row=r, column=sub_col).value or "").strip() or None if sub_col else None
        person = str(ws.cell(row=r, column=person_col).value or "").strip() or None if person_col else None
        
        row_hash = sha256_text(
            "|".join([
                "debit",
                txn_dt.isoformat(),
                f"{amount:.2f}",
                description,
            ])
        )
        
        rows.append({
            "row_hash": row_hash,
            "txn_date": txn_dt.isoformat(),
            "cash_date": txn_dt.isoformat(),
            "amount": amount,
            "description": description,
            "group_name": None,
            "category": category,
            "subcategory": subcategory,
            "payment_method": "debit",
            "account": None,
            "source": "unified_xlsx",
            "statement_closing_date": None,
            "statement_due_date": None,
            "person": person,
            "reimbursable": 0,
            "reference": None,
            "notes": None,
            "source_file": source_file,
            "source_hash": source_hash,
            "external_id": None,
            "created_at": created,
            "updated_at": created,
        })
    
    return rows


def _import_income_sheet(
    ws, source_hash: str, source_file: str, created: str
) -> list[dict[str, Any]]:
    """Importa aba Receitas"""
    headers = _get_headers(ws)
    
    date_col = _find_col(headers, ("data", "date"))
    desc_col = _find_col(headers, ("descricao", "descrição", "description"))
    amount_col = _find_col(headers, ("valor", "amount"))
    cat_col = _find_col(headers, ("categoria", "category"))
    person_col = _find_col(headers, ("pessoa", "person"))
    
    if not all([date_col, desc_col, amount_col]):
        return []
    
    rows: list[dict[str, Any]] = []
    
    for r in range(2, ws.max_row + 1):
        txn_dt = parse_date(ws.cell(row=r, column=date_col).value) if date_col else None
        if txn_dt is None:
            continue
        
        description = str(ws.cell(row=r, column=desc_col).value or "").strip()
        if not description:
            continue
        
        amount_raw = parse_brl_number(ws.cell(row=r, column=amount_col).value)
        if amount_raw is None:
            continue
        amount = abs(float(amount_raw))  # Income is positive
        
        category = str(ws.cell(row=r, column=cat_col).value or "").strip() or None if cat_col else None
        person = str(ws.cell(row=r, column=person_col).value or "").strip() or None if person_col else None
        
        row_hash = sha256_text(
            "|".join([
                "income",
                txn_dt.isoformat(),
                f"{amount:.2f}",
                description,
            ])
        )
        
        rows.append({
            "row_hash": row_hash,
            "txn_date": txn_dt.isoformat(),
            "cash_date": txn_dt.isoformat(),
            "amount": amount,
            "description": description,
            "group_name": None,
            "category": category,
            "subcategory": None,
            "payment_method": "income",
            "account": None,
            "source": "unified_xlsx",
            "statement_closing_date": None,
            "statement_due_date": None,
            "person": person,
            "reimbursable": 0,
            "reference": None,
            "notes": None,
            "source_file": source_file,
            "source_hash": source_hash,
            "external_id": None,
            "created_at": created,
            "updated_at": created,
        })
    
    return rows


def _import_household_sheet(
    ws, source_hash: str, source_file: str, created: str
) -> list[dict[str, Any]]:
    """
    Importa aba Contas da Casa.
    Essas são contas fixas divididas no início do mês corrente.
    Usa payment_method = 'household' para distinguir na reconciliação.
    """
    headers = _get_headers(ws)
    
    ref_month_col = _find_col(headers, ("mes de referencia", "mês de referência", "mes referencia"))
    cat_col = _find_col(headers, ("categoria", "category"))
    subcat_col = _find_col(headers, ("subcategoria", "subcategory", "sub"))
    desc_col = _find_col(headers, ("descricao", "descrição", "description"))
    amount_col = _find_col(headers, ("valor", "amount"))
    date_col = _find_col(headers, ("data de pagamento", "data pagamento", "data"))
    person_col = _find_col(headers, ("pago por", "quem pagou", "pessoa", "person"))
    
    if not all([amount_col]):
        return []
    
    rows: list[dict[str, Any]] = []
    
    for r in range(2, ws.max_row + 1):
        amount_raw = parse_brl_number(ws.cell(row=r, column=amount_col).value)
        if amount_raw is None:
            continue
        amount = -abs(float(amount_raw))
        
        # Reference month - expected format: "2026-01" or similar
        ref_month = str(ws.cell(row=r, column=ref_month_col).value or "").strip() if ref_month_col else None
        
        category = str(ws.cell(row=r, column=cat_col).value or "").strip() if cat_col else None
        subcategory = str(ws.cell(row=r, column=subcat_col).value or "").strip() if subcat_col else None
        description = str(ws.cell(row=r, column=desc_col).value or "").strip() if desc_col else category or "(conta da casa)"
        
        txn_dt = parse_date(ws.cell(row=r, column=date_col).value) if date_col else None
        person = str(ws.cell(row=r, column=person_col).value or "").strip() or None if person_col else None
        
        # Use reference month as the transaction date if no specific date
        if txn_dt is None and ref_month:
            # Try to parse "2026-01" as first day of month
            try:
                txn_dt = parse_date(f"{ref_month}-01")
            except Exception:
                pass
        
        if txn_dt is None:
            continue
        
        row_hash = sha256_text(
            "|".join([
                "household",
                txn_dt.isoformat(),
                f"{amount:.2f}",
                description,
                category or "",
                subcategory or "",
            ])
        )
        
        rows.append({
            "row_hash": row_hash,
            "txn_date": txn_dt.isoformat(),
            "cash_date": txn_dt.isoformat(),
            "amount": amount,
            "description": description,
            "group_name": None,
            "category": category,
            "subcategory": subcategory,
            "payment_method": "household",  # Special type for reconciliation
            "account": None,
            "source": "unified_xlsx",
            "statement_closing_date": None,
            "statement_due_date": None,
            "person": person,
            "reimbursable": 0,
            "reference": ref_month,  # Store reference month here
            "notes": None,
            "source_file": source_file,
            "source_hash": source_hash,
            "external_id": None,
            "created_at": created,
            "updated_at": created,
        })
    
    return rows

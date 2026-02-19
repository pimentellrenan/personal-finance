from __future__ import annotations

import io
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd

from pf import queries as pf_queries
from pf.templates import build_debit_template_bytes, build_income_template_bytes


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(
            "Dependência ausente: openpyxl. Instale com `pip install -r requirements.txt`."
        ) from e


def _atomic_write_bytes(path: Path, content: bytes) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_bytes(content)
    tmp.replace(path)


def _write_debit_workbook(path: Path, *, df_all: pd.DataFrame, expense_categories_tree: dict[str, Any]) -> None:
    _require_openpyxl()
    from openpyxl import load_workbook

    buf = io.BytesIO(build_debit_template_bytes(expense_categories_tree=expense_categories_tree))
    wb = load_workbook(buf)
    ws = wb["Dados"]
    ws_lists = wb["Listas"]

    df = df_all.copy()
    if df.empty:
        out_rows = []
    else:
        df = df[(df["amount"] < 0) & (df["payment_method"].isin(["debit", "pix", "transfer", "cash"]))]
        df = df.sort_values(["txn_date", "id"], ascending=[True, True])

        out_rows = []
        lookup_end = 1
        for i in range(1, ws_lists.max_row + 1):
            if not ws_lists.cell(row=i, column=1).value:
                break
            lookup_end = i
        for _, r in df.iterrows():
            grp = str(r.get("group_name") or "").strip()
            cat = str(r.get("category") or "").strip()
            sub = str(r.get("subcategory") or "").strip()
            if sub:
                category = cat
                subcategory = sub
            elif grp and cat:
                category = grp
                subcategory = cat
            else:
                category = cat or grp
                subcategory = sub

            row_num = 2 + len(out_rows)
            out_rows.append(
                [
                    r.get("txn_date"),
                    r.get("txn_date"),
                    category,
                    subcategory,
                    "Débito",
                    r.get("description"),
                    float(abs(r.get("amount") or 0.0)),
                    "Pago",
                    f'=IFERROR(VLOOKUP($C{row_num},Listas!$A$1:$B${lookup_end},2,FALSE),"")',
                ]
            )

    # Clear existing rows (keep header)
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row in out_rows:
        ws.append(row)

    out = io.BytesIO()
    wb.save(out)
    _atomic_write_bytes(path, out.getvalue())


def _write_income_workbook(path: Path, *, df_all: pd.DataFrame, income_categories_tree: dict[str, Any]) -> None:
    _require_openpyxl()
    from openpyxl import load_workbook

    buf = io.BytesIO(build_income_template_bytes(income_categories_tree=income_categories_tree))
    wb = load_workbook(buf)
    ws = wb["Dados"]

    df = df_all.copy()
    if df.empty:
        out_rows = []
    else:
        df = df[df["payment_method"] == "income"]
        df = df.sort_values(["txn_date", "id"], ascending=[True, True])

        out_rows = []
        for _, r in df.iterrows():
            out_rows.append(
                [
                    r.get("txn_date"),
                    r.get("description"),
                    float(abs(r.get("amount") or 0.0)),
                    r.get("category") or "",
                    r.get("subcategory") or "",
                ]
            )

    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in out_rows:
        ws.append(row)

    out = io.BytesIO()
    wb.save(out)
    _atomic_write_bytes(path, out.getvalue())


def _write_all_workbook(path: Path, *, df_all: pd.DataFrame) -> None:
    _require_openpyxl()
    df = df_all.copy()
    if df.empty:
        df = pd.DataFrame(
            columns=[
                "id",
                "txn_date",
                "cash_date",
                "amount",
                "description",
                "payment_method",
                "account",
                "category",
                "subcategory",
                "reimbursable",
                "reference",
                "notes",
                "statement_closing_date",
                "statement_due_date",
            ]
        )
    df = df.sort_values(["cash_date", "id"], ascending=[True, True])
    cols = [
        "id",
        "txn_date",
        "cash_date",
        "amount",
        "description",
        "payment_method",
        "account",
        "category",
        "subcategory",
        "reimbursable",
        "reference",
        "notes",
        "statement_closing_date",
        "statement_due_date",
    ]
    cols = [c for c in cols if c in df.columns]
    df = df[cols]

    tmp = path.with_suffix(path.suffix + ".tmp")
    with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Lançamentos")
    tmp.replace(path)


def _write_credit_card_workbook(path: Path, *, df_all: pd.DataFrame, expense_categories_tree: dict[str, Any] | None = None) -> None:
    """
    Gera/atualiza o Excel de cartão de crédito.
    
    Se o Excel já existe: PRESERVA os dados existentes e ADICIONA novos registros do banco.
    Se não existe: cria um novo Excel com todos os dados do banco.
    """
    _require_openpyxl()
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    expected_cols = [
        "row_hash",
        "txn_date",
        "statement_due_date",
        "category",
        "subcategory",
        "account",
        "description",
        "amount",
    ]

    # Pegar dados do banco (só credit_card)
    df = df_all.copy()
    if df.empty or "payment_method" not in df.columns:
        df = pd.DataFrame(columns=expected_cols)
    else:
        df = df[df["payment_method"] == "credit_card"]
        if df.empty:
            df = pd.DataFrame(columns=expected_cols)

    if not df.empty:
        sort_cols = [c for c in ["statement_due_date", "txn_date", "id"] if c in df.columns]
        if sort_cols:
            df = df.sort_values(sort_cols, ascending=[True] * len(sort_cols))

    cols = [
        "row_hash",
        "txn_date",
        "statement_due_date",
        "category",
        "subcategory",
        "account",
        "description",
        "amount",
    ]
    cols = [c for c in cols if c in df.columns]
    df_from_db = df[cols].copy()

    # Se o Excel já existe, ler os dados atuais e mesclar
    if path.exists():
        try:
            df_existing = pd.read_excel(path, sheet_name=0)
            
            # Renomear colunas PT → EN para merge
            col_map = {
                "Hash (oculto)": "row_hash",
                "Data da compra": "txn_date",
                "Data do vencimento": "statement_due_date",
                "Categoria": "category",
                "Subcategoria": "subcategory",
                "Cartão de crédito": "account",
                "Descrição": "description",
                "Valor (R$)": "amount",
            }
            df_existing = df_existing.rename(columns=col_map)
            
            # Limpar e converter tipos
            if "row_hash" in df_existing.columns:
                df_existing["row_hash"] = df_existing["row_hash"].fillna("").astype(str).str.strip()
                # Remover linhas com hash vazio (linhas vazias do Excel)
                df_existing = df_existing[df_existing["row_hash"] != ""]
            
            # Pegar hashes que já existem no Excel
            existing_hashes = set(df_existing["row_hash"].tolist()) if "row_hash" in df_existing.columns else set()
            
            # Novos registros do banco que NÃO estão no Excel
            if "row_hash" in df_from_db.columns:
                df_new_from_db = df_from_db[~df_from_db["row_hash"].isin(existing_hashes)].copy()
            else:
                df_new_from_db = pd.DataFrame(columns=expected_cols)
            
            # MESCLAR: dados do Excel (preservados) + novos do banco
            df_merged = pd.concat([df_existing[cols], df_new_from_db], ignore_index=True)
            
            # Re-ordenar por data
            if "statement_due_date" in df_merged.columns and "txn_date" in df_merged.columns:
                df_merged = df_merged.sort_values(["statement_due_date", "txn_date"], ascending=[True, True])
            
            df = df_merged
        except Exception:
            # Se falhar ao ler Excel existente, usar apenas dados do banco
            df = df_from_db
    else:
        # Excel não existe, usar dados do banco
        df = df_from_db

    today = pd.Timestamp(date.today()).date()
    if "statement_due_date" in df.columns:
        df["status"] = df["statement_due_date"].apply(
            lambda d: "Pago" if pd.notna(d) and d <= today else "Em aberto"
        )
    else:
        df["status"] = "Em aberto"

    # Friendly PT-BR headers and fewer date columns (keep only purchase date + due date).
    df = df.rename(
        columns={
            "row_hash": "Hash (oculto)",
            "txn_date": "Data da compra",
            "statement_due_date": "Data do vencimento",
            "category": "Categoria",
            "subcategory": "Subcategoria",
            "account": "Cartão de crédito",
            "description": "Descrição",
            "amount": "Valor (R$)",
            "status": "Status (cartao pago ou nao)",
        }
    )

    # Replace NaN/None/"nan"/"None" with empty strings for text columns
    text_cols = ["Categoria", "Subcategoria", "Descrição", "Cartão de crédito", "Hash (oculto)", "Status (cartao pago ou nao)"]
    for col in text_cols:
        if col in df.columns:
            # Handle various forms of "empty": NaN, None, "nan", "None", ""
            df[col] = df[col].apply(
                lambda v: "" if (pd.isna(v) or v is None or str(v).lower() in ("nan", "none", "")) else str(v)
            )

    # Use .tmp.xlsx instead of .xlsx.tmp so openpyxl can read it
    tmp = path.parent / (path.stem + ".tmp.xlsx")
    with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Cartão", na_rep="")

    wb = load_workbook(tmp)
    ws = wb["Cartão"]
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].hidden = True
    
    # Add category dropdowns if categories tree is provided
    if expense_categories_tree:
        categories = list(expense_categories_tree.keys())
        
        # Create a hidden sheet for category lists
        if "_Listas" in wb.sheetnames:
            del wb["_Listas"]
        ws_lists = wb.create_sheet("_Listas")
        ws_lists.sheet_state = "hidden"
        
        # Write categories in column A
        ws_lists["A1"] = "Categorias"
        for i, cat in enumerate(categories, start=2):
            ws_lists[f"A{i}"] = cat
        
        # Create named range for categories list
        from openpyxl.workbook.defined_name import DefinedName
        cat_range_ref = f"'_Listas'!$A$2:$A${len(categories) + 1}"
        if "Categorias" in wb.defined_names:
            del wb.defined_names["Categorias"]
        wb.defined_names["Categorias"] = DefinedName("Categorias", attr_text=cat_range_ref)
        
        # Write subcategories for each category in subsequent columns
        # Create named ranges for INDIRECT formula
        col_idx = 2  # Start at column B
        for cat in categories:
            subs = expense_categories_tree.get(cat, [])
            if isinstance(subs, list) and subs:
                col_letter = get_column_letter(col_idx)
                # Clean category name for use as range name (remove special chars, accents)
                import unicodedata
                safe_cat = unicodedata.normalize('NFD', cat)
                safe_cat = ''.join(c for c in safe_cat if unicodedata.category(c) != 'Mn')
                safe_cat = "".join(c if c.isalnum() else "_" for c in safe_cat)
                # Named range must start with letter or underscore
                if safe_cat and safe_cat[0].isdigit():
                    safe_cat = "_" + safe_cat
                
                ws_lists[f"{col_letter}1"] = cat
                for j, sub in enumerate(subs, start=2):
                    ws_lists[f"{col_letter}{j}"] = sub
                
                # Create named range for this category's subcategories
                range_ref = f"'_Listas'!${col_letter}$2:${col_letter}${len(subs) + 1}"
                if safe_cat in wb.defined_names:
                    del wb.defined_names[safe_cat]
                defn = DefinedName(safe_cat, attr_text=range_ref)
                wb.defined_names[safe_cat] = defn
                
                col_idx += 1
        
        # Determine how many rows of data we have
        max_row = max(ws.max_row, 100) if ws.max_row else 1000
        
        # Add data validation for Categoria column (column D = 4)
        dv_cat = DataValidation(
            type="list",
            formula1="Categorias",
            allow_blank=True,
            showDropDown=False,  # False means show the dropdown arrow
        )
        dv_cat.error = "Selecione uma categoria válida"
        dv_cat.errorTitle = "Categoria inválida"
        dv_cat.prompt = "Selecione a categoria"
        dv_cat.promptTitle = "Categoria"
        ws.add_data_validation(dv_cat)
        dv_cat.add(f"D2:D{max_row}")
        
        # Add data validation for Subcategoria using INDIRECT to filter by Categoria
        # The formula references the named range that matches the category name
        # We need to apply validation row by row for INDIRECT to work correctly
        for row in range(2, max_row + 1):
            # Create a formula that converts the category name to the named range
            # We sanitize the category name the same way we created the named ranges
            # Using SUBSTITUTE to handle special characters
            indirect_formula = f'INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(D{row}," ","_"),"ã","a"),"á","a"),"é","e"),"í","i"),"ç","c"))'
            
            dv_sub = DataValidation(
                type="list",
                formula1=indirect_formula,
                allow_blank=True,
                showDropDown=False,
            )
            dv_sub.error = "Selecione uma subcategoria válida"
            dv_sub.errorTitle = "Subcategoria inválida"
            ws.add_data_validation(dv_sub)
            dv_sub.add(f"E{row}")
    
    wb.save(tmp)
    tmp.replace(path)


def sync_excel_files(
    base_dir: Path,
    *,
    conn,
    expense_categories_tree: dict[str, Any],
    income_categories_tree: dict[str, Any],
    include_debit: bool = True,
    include_income: bool = True,
    include_credit_card: bool = True,
) -> dict[str, Path]:
    templates_dir = base_dir / "templates"
    templates_dir.mkdir(parents=True, exist_ok=True)

    debit_path = templates_dir / "saida_debitos.xlsx"
    income_path = templates_dir / "saida_receitas.xlsx"
    credit_card_path = templates_dir / "cartao_credito.xlsx"

    df_all = pf_queries.load_transactions_df(conn)

    has_debit = (
        (not df_all.empty)
        and ("payment_method" in df_all.columns)
        and ("amount" in df_all.columns)
        and ((df_all["amount"] < 0) & (df_all["payment_method"].isin(["debit", "pix", "transfer", "cash"]))).any()
    )
    has_income = (
        (not df_all.empty)
        and ("payment_method" in df_all.columns)
        and (df_all["payment_method"] == "income").any()
    )

    if include_debit and has_debit:
        _write_debit_workbook(debit_path, df_all=df_all, expense_categories_tree=expense_categories_tree)
    if include_income and has_income:
        _write_income_workbook(income_path, df_all=df_all, income_categories_tree=income_categories_tree)
    # Always generate credit card workbook when requested (even if empty)
    if include_credit_card:
        _write_credit_card_workbook(credit_card_path, df_all=df_all, expense_categories_tree=expense_categories_tree)

    return {"debit": debit_path, "income": income_path, "credit_card": credit_card_path}
